#!/usr/bin/env python3
"""Generate initial benchmark fixtures and gold manifests from repository sample files."""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

import xlrd
from openpyxl import load_workbook


BENCHMARK_SLUG = "search-docs-of-crs-agent"
IMAGE_PREPROCESS_STRATEGY = "ocr_then_context_injection"
COMPONENT_TEXT_CASE_IDS = ("case_000003",)
COMPONENT_IMAGE_CASE_IDS = ("case_000002", "case_000007", "case_000010")
VISIBLE_ACCEPTANCE_CASE_IDS = ("case_000002", "case_000003", "case_000007", "case_000008", "case_000010")
BLIND_ACCEPTANCE_CASE_IDS = ("case_000004", "case_000009")
SCORING_TOP_K = 10

NOISE_QUERIES = (
    "老师，帮我找下火星牌 MARS-42 BCM 针脚图 ZZZ_NO_SUCH_DOC_001",
    "宇宙牌 QX999 整车电路图 ZZZ_NO_SUCH_DOC_002",
    "北极星 PX-700 量子 ECU 电路图 ZZZ_NO_SUCH_DOC_003",
    "海王星 N7 冷藏车仪表定义图 ZZZ_NO_SUCH_DOC_004",
    "虚构品牌 TEST-ALPHA 发动机线束图 ZZZ_NO_SUCH_DOC_005",
    "不存在车型 OMEGA-9000 电器原理图 ZZZ_NO_SUCH_DOC_006",
)


def repo_root() -> Path:
    return Path(__file__).resolve().parents[3]


def benchmark_root() -> Path:
    return Path(__file__).resolve().parents[1]


def sample_root() -> Path:
    return repo_root() / "sample"


def fixtures_root() -> Path:
    return benchmark_root() / "fixtures"


def gold_root() -> Path:
    return benchmark_root() / "gold"


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def split_image_tokens(raw: str | None) -> list[str]:
    if not raw:
        return []
    return [item for item in re.split(r"[;\s]+", str(raw).strip()) if item]


def split_gold_titles(raw: str | None) -> list[str]:
    if not raw:
        return []
    titles = []
    for item in re.split(r"(?:\r?\n)+|、", str(raw)):
        cleaned = str(item).strip()
        if cleaned:
            titles.append(cleaned)
    return titles


def split_keyword_aliases(raw: str | None) -> list[str]:
    if not raw:
        return []
    seen: set[str] = set()
    output: list[str] = []
    for item in re.split(r"[,，]+", str(raw)):
        cleaned = str(item).strip()
        if cleaned and cleaned not in seen:
            seen.add(cleaned)
            output.append(cleaned)
    return output


def relativize(path: Path) -> str:
    return path.relative_to(repo_root()).as_posix()


def resolve_question_images(raw: str | None, sample_files: dict[str, Path]) -> list[str]:
    images: list[str] = []
    for token in split_image_tokens(raw):
        filename = Path(token).name
        if filename in sample_files:
            images.append(relativize(sample_files[filename]))
        else:
            images.append(token.replace("\\", "/"))
    return images


def build_suite_payload(
    *,
    suite_id: str,
    layer: str,
    source_files: list[str],
    cases: list[dict[str, Any]],
) -> dict[str, Any]:
    return {
        "benchmark_slug": BENCHMARK_SLUG,
        "suite_id": suite_id,
        "layer": layer,
        "source_files": source_files,
        "case_count": len(cases),
        "cases": cases,
    }


def build_gold_payload(
    *,
    suite_id: str,
    layer: str,
    source_files: list[str],
    cases: list[dict[str, Any]],
    acceptance_threshold: float,
) -> dict[str, Any]:
    return {
        "benchmark_slug": BENCHMARK_SLUG,
        "suite_id": suite_id,
        "layer": layer,
        "source_files": source_files,
        "acceptance_threshold": acceptance_threshold,
        "case_count": len(cases),
        "cases": cases,
    }


def build_base_fixture_case(
    *,
    case_id: str,
    layer: str,
    input_modality: str,
    question_text: str,
    question_images: list[str],
    vehicle_info: str | None,
    benchmark_track: str,
    notes: str | None = None,
    extra_fields: dict[str, Any] | None = None,
) -> dict[str, Any]:
    case = {
        "case_id": case_id,
        "layer": layer,
        "input_modality": input_modality,
        "question_text": question_text,
        "question_images": question_images,
        "vehicle_info": vehicle_info,
        "benchmark_track": benchmark_track,
        "preprocess_strategy": IMAGE_PREPROCESS_STRATEGY if input_modality == "image_text" else "none",
        "request_context": {},
        "notes": notes,
    }
    if extra_fields:
        case.update(extra_fields)
    return case


def build_base_gold_case(
    *,
    case_id: str,
    layer: str,
    accepted_titles: list[str],
    preferred_title: str | None,
    expected_response_type: str = "documents",
    top_k: int = SCORING_TOP_K,
    extra_fields: dict[str, Any] | None = None,
) -> dict[str, Any]:
    case = {
        "case_id": case_id,
        "layer": layer,
        "accepted_titles": accepted_titles,
        "preferred_title": preferred_title,
        "expected_response_type": expected_response_type,
        "top_k": top_k,
    }
    if extra_fields:
        case.update(extra_fields)
    return case


def load_real_cases(sample_files: dict[str, Path]) -> dict[str, dict[str, Any]]:
    workbook = load_workbook(sample_root() / "benchmark_excel_template.xlsx", read_only=True, data_only=True)
    sheet = workbook.active
    rows: dict[str, dict[str, Any]] = {}
    for row in sheet.iter_rows(min_row=5, values_only=True):
        if not row or not row[0]:
            continue
        case_id = str(row[0]).strip()
        if not case_id or case_id == "case_000001":
            continue
        gold_titles = split_gold_titles(row[5])
        rows[case_id] = {
            "case_id": case_id,
            "question_text": str(row[1] or "").strip(),
            "question_images": resolve_question_images(row[2], sample_files),
            "vehicle_info": str(row[3]).strip() if row[3] else None,
            "teacher_reply": str(row[4]).strip() if row[4] else None,
            "accepted_titles": gold_titles,
            "preferred_title": gold_titles[0] if gold_titles else None,
            "question_type": str(row[6]).strip() if row[6] else None,
            "remark": str(row[7]).strip() if row[7] else None,
        }
    return rows


def find_mock_source_files() -> tuple[Path, Path]:
    csv_candidates = sorted(path for path in sample_root().iterdir() if path.suffix.lower() == ".csv")
    txt_candidates = sorted(path for path in sample_root().iterdir() if path.suffix.lower() == ".txt")
    if len(csv_candidates) != 1 or len(txt_candidates) != 1:
        raise RuntimeError("missing mock source files")
    return csv_candidates[0], txt_candidates[0]


def load_mock_rows(source_path: Path, *, keyword_header: str) -> list[dict[str, Any]]:
    workbook = xlrd.open_workbook(file_contents=source_path.read_bytes())
    sheet = workbook.sheet_by_index(0)
    header = [str(value).strip() for value in sheet.row_values(0)]
    try:
        title_idx = header.index("关联文件名称")
        keyword_idx = header.index(keyword_header)
    except ValueError as exc:
        raise RuntimeError(f"missing expected columns in {source_path.name}") from exc

    rows: list[dict[str, Any]] = []
    for row_idx in range(1, sheet.nrows):
        title = str(sheet.cell_value(row_idx, title_idx)).strip()
        keywords = str(sheet.cell_value(row_idx, keyword_idx)).strip()
        if not title or not keywords:
            continue
        aliases = split_keyword_aliases(keywords)
        if not aliases:
            continue
        rows.append(
            {
                "title": title,
                "primary_query": aliases[0],
                "alternate_queries": aliases[1:],
            }
        )
    return rows


def build_mock_suite_rows(
    *,
    rows: list[dict[str, Any]],
    case_prefix: str,
    source_file: str,
    notes: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    fixture_cases: list[dict[str, Any]] = []
    gold_cases: list[dict[str, Any]] = []
    for idx, row in enumerate(rows, start=1):
        case_id = f"{case_prefix}_{idx:04d}"
        fixture_cases.append(
            build_base_fixture_case(
                case_id=case_id,
                layer="atomic",
                input_modality="text",
                question_text=row["primary_query"],
                question_images=[],
                vehicle_info=None,
                benchmark_track="search_api",
                notes=notes,
                extra_fields={
                    "alternate_queries": row["alternate_queries"],
                    "source_file": source_file,
                },
            )
        )
        gold_cases.append(
            build_base_gold_case(
                case_id=case_id,
                layer="atomic",
                accepted_titles=[row["title"]],
                preferred_title=row["title"],
                extra_fields={"source_file": source_file},
            )
        )
    return fixture_cases, gold_cases


def build_noise_suite_rows() -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    fixture_cases: list[dict[str, Any]] = []
    gold_cases: list[dict[str, Any]] = []
    for idx, query in enumerate(NOISE_QUERIES, start=1):
        case_id = f"noise_{idx:04d}"
        fixture_cases.append(
            build_base_fixture_case(
                case_id=case_id,
                layer="atomic",
                input_modality="text",
                question_text=query,
                question_images=[],
                vehicle_info=None,
                benchmark_track="search_api",
                notes="synthetic negative query for false-positive suppression",
            )
        )
        gold_cases.append(
            build_base_gold_case(
                case_id=case_id,
                layer="atomic",
                accepted_titles=[],
                preferred_title=None,
                expected_response_type="message_or_empty",
            )
        )
    return fixture_cases, gold_cases


def build_real_fixture_case(case: dict[str, Any], layer: str) -> dict[str, Any]:
    input_modality = "image_text" if case["question_images"] else "text"
    notes = case["remark"]
    if input_modality == "image_text":
        image_note = "图文 case 需先完成图片识别并把结果注入 ChatRequest.context；当前 fixture 先保留空 request_context。"
        notes = image_note if not notes else f"{notes}；{image_note}"
    return build_base_fixture_case(
        case_id=case["case_id"],
        layer=layer,
        input_modality=input_modality,
        question_text=case["question_text"],
        question_images=case["question_images"],
        vehicle_info=case["vehicle_info"],
        benchmark_track="chat_completions",
        notes=notes,
        extra_fields={
            "teacher_reply": case["teacher_reply"],
            "question_type": case["question_type"],
        },
    )


def build_real_gold_case(case: dict[str, Any], layer: str) -> dict[str, Any]:
    return build_base_gold_case(
        case_id=case["case_id"],
        layer=layer,
        accepted_titles=list(case["accepted_titles"]),
        preferred_title=case["preferred_title"],
        extra_fields={"question_type": case["question_type"]},
    )


def select_cases(real_cases: dict[str, dict[str, Any]], case_ids: tuple[str, ...]) -> list[dict[str, Any]]:
    return [real_cases[case_id] for case_id in case_ids]


def sample_file_lookup() -> dict[str, Path]:
    return {path.name: path for path in sample_root().iterdir() if path.is_file()}


def main() -> int:
    sample_files = sample_file_lookup()
    real_cases = load_real_cases(sample_files)
    dongfeng_file, jiefang_file = find_mock_source_files()
    dongfeng_rows = load_mock_rows(dongfeng_file, keyword_header="关键词")
    jiefang_rows = load_mock_rows(jiefang_file, keyword_header="搜索可能关键字")

    dongfeng_fixture_cases, dongfeng_gold_cases = build_mock_suite_rows(
        rows=dongfeng_rows,
        case_prefix="mock_dongfeng",
        source_file=relativize(dongfeng_file),
        notes="mock 东风关键词到关联文件名称的原子召回样本",
    )
    jiefang_fixture_cases, jiefang_gold_cases = build_mock_suite_rows(
        rows=jiefang_rows,
        case_prefix="mock_jiefang",
        source_file=relativize(jiefang_file),
        notes="mock 解放关键词到关联文件名称的原子召回样本",
    )
    noise_fixture_cases, noise_gold_cases = build_noise_suite_rows()

    component_text_cases = select_cases(real_cases, COMPONENT_TEXT_CASE_IDS)
    component_image_cases = select_cases(real_cases, COMPONENT_IMAGE_CASE_IDS)
    visible_cases = select_cases(real_cases, VISIBLE_ACCEPTANCE_CASE_IDS)
    blind_cases = select_cases(real_cases, BLIND_ACCEPTANCE_CASE_IDS)

    write_json(
        fixtures_root() / "01_atomic" / "mock_dongfeng_keyword_recall.json",
        build_suite_payload(
            suite_id="mock_dongfeng_keyword_recall",
            layer="atomic",
            source_files=[relativize(dongfeng_file)],
            cases=dongfeng_fixture_cases,
        ),
    )
    write_json(
        gold_root() / "01_atomic" / "mock_dongfeng_keyword_recall.json",
        build_gold_payload(
            suite_id="mock_dongfeng_keyword_recall",
            layer="atomic",
            source_files=[relativize(dongfeng_file)],
            cases=dongfeng_gold_cases,
            acceptance_threshold=0.85,
        ),
    )

    write_json(
        fixtures_root() / "01_atomic" / "mock_jiefang_keyword_recall.json",
        build_suite_payload(
            suite_id="mock_jiefang_keyword_recall",
            layer="atomic",
            source_files=[relativize(jiefang_file)],
            cases=jiefang_fixture_cases,
        ),
    )
    write_json(
        gold_root() / "01_atomic" / "mock_jiefang_keyword_recall.json",
        build_gold_payload(
            suite_id="mock_jiefang_keyword_recall",
            layer="atomic",
            source_files=[relativize(jiefang_file)],
            cases=jiefang_gold_cases,
            acceptance_threshold=0.85,
        ),
    )

    write_json(
        fixtures_root() / "01_atomic" / "noise" / "synthetic_noise_queries.json",
        build_suite_payload(
            suite_id="synthetic_noise_queries",
            layer="atomic",
            source_files=[],
            cases=noise_fixture_cases,
        ),
    )
    write_json(
        gold_root() / "01_atomic" / "noise" / "synthetic_noise_queries.json",
        build_gold_payload(
            suite_id="synthetic_noise_queries",
            layer="atomic",
            source_files=[],
            cases=noise_gold_cases,
            acceptance_threshold=1.0,
        ),
    )

    write_json(
        fixtures_root() / "02_component" / "real_text_single_turn.json",
        build_suite_payload(
            suite_id="real_text_single_turn",
            layer="component",
            source_files=[relativize(sample_root() / "benchmark_excel_template.xlsx")],
            cases=[build_real_fixture_case(case, "component") for case in component_text_cases],
        ),
    )
    write_json(
        gold_root() / "02_component" / "real_text_single_turn.json",
        build_gold_payload(
            suite_id="real_text_single_turn",
            layer="component",
            source_files=[relativize(sample_root() / "benchmark_excel_template.xlsx")],
            cases=[build_real_gold_case(case, "component") for case in component_text_cases],
            acceptance_threshold=1.0,
        ),
    )

    write_json(
        fixtures_root() / "02_component" / "real_image_augmented_single_turn.json",
        build_suite_payload(
            suite_id="real_image_augmented_single_turn",
            layer="component",
            source_files=[relativize(sample_root() / "benchmark_excel_template.xlsx")],
            cases=[build_real_fixture_case(case, "component") for case in component_image_cases],
        ),
    )
    write_json(
        gold_root() / "02_component" / "real_image_augmented_single_turn.json",
        build_gold_payload(
            suite_id="real_image_augmented_single_turn",
            layer="component",
            source_files=[relativize(sample_root() / "benchmark_excel_template.xlsx")],
            cases=[build_real_gold_case(case, "component") for case in component_image_cases],
            acceptance_threshold=1.0,
        ),
    )

    write_json(
        fixtures_root() / "03_e2e" / "real_acceptance_visible.json",
        build_suite_payload(
            suite_id="real_acceptance_visible",
            layer="e2e",
            source_files=[relativize(sample_root() / "benchmark_excel_template.xlsx")],
            cases=[build_real_fixture_case(case, "e2e") for case in visible_cases],
        ),
    )
    write_json(
        gold_root() / "03_e2e" / "real_acceptance_visible.json",
        build_gold_payload(
            suite_id="real_acceptance_visible",
            layer="e2e",
            source_files=[relativize(sample_root() / "benchmark_excel_template.xlsx")],
            cases=[build_real_gold_case(case, "e2e") for case in visible_cases],
            acceptance_threshold=0.85,
        ),
    )

    write_json(
        fixtures_root() / "04_blind" / "real_acceptance_holdout.json",
        build_suite_payload(
            suite_id="real_acceptance_holdout",
            layer="blind",
            source_files=[relativize(sample_root() / "benchmark_excel_template.xlsx")],
            cases=[build_real_fixture_case(case, "blind") for case in blind_cases],
        ),
    )
    write_json(
        gold_root() / "04_blind" / "real_acceptance_holdout.json",
        build_gold_payload(
            suite_id="real_acceptance_holdout",
            layer="blind",
            source_files=[relativize(sample_root() / "benchmark_excel_template.xlsx")],
            cases=[build_real_gold_case(case, "blind") for case in blind_cases],
            acceptance_threshold=0.85,
        ),
    )

    print("[OK] generated benchmark fixtures and gold manifests")
    print(f" - mock_dongfeng: {len(dongfeng_fixture_cases)}")
    print(f" - mock_jiefang: {len(jiefang_fixture_cases)}")
    print(f" - noise: {len(noise_fixture_cases)}")
    print(f" - component_text: {len(component_text_cases)}")
    print(f" - component_image: {len(component_image_cases)}")
    print(f" - e2e_visible: {len(visible_cases)}")
    print(f" - blind_holdout: {len(blind_cases)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
