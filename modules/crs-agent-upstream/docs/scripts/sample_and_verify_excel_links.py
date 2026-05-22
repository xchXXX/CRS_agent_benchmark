#!/usr/bin/env python3
"""Sample candidate rows from the master sheet and verify link types."""

from __future__ import annotations

import argparse
import base64
import json
import os
import random
import time
import unicodedata
from pathlib import Path
from typing import Any

import httpx
from openpyxl import Workbook, load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT = SCRIPT_DIR / "完整的电路图的目录.xlsx"
DEFAULT_OUTPUT = SCRIPT_DIR / "完整的电路图的目录_随机350_校验300.xlsx"
DEFAULT_CANDIDATE_SIZE = 350
DEFAULT_TARGET_SIZE = 300
DEFAULT_DELAY_SECONDS = 0.8
DEFAULT_PAGE_SIZE = 5
SEARCH_URL = "https://wx.51gonggui.com/commonrail/api/management/getMeansList.json"
VALIDATE_URL = "https://wx.51gonggui.com/commonrail/api/member-api/userLoginInfo.json"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="从主表随机抽样并校验外部搜索首条链接类型。")
    parser.add_argument("--input", default=str(DEFAULT_INPUT), help="主表 xlsx 路径")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT), help="输出 xlsx 路径")
    parser.add_argument("--sheet", default=None, help="源工作表名称；默认使用第一个工作表")
    parser.add_argument("--token-env", default="GGZJ_APP_TOKEN", help="读取 token 的环境变量名")
    parser.add_argument(
        "--candidate-size",
        type=int,
        default=DEFAULT_CANDIDATE_SIZE,
        help="先随机抽取的候选记录数",
    )
    parser.add_argument(
        "--target-size",
        type=int,
        default=DEFAULT_TARGET_SIZE,
        help="最终保留的有效记录数目标",
    )
    parser.add_argument("--seed", type=int, default=None, help="随机种子；指定后可复现抽样结果")
    parser.add_argument(
        "--delay",
        type=float,
        default=DEFAULT_DELAY_SECONDS,
        help="每次搜索后的等待秒数，避免请求过快",
    )
    parser.add_argument("--page-size", type=int, default=DEFAULT_PAGE_SIZE, help="每次搜索拉取的结果数")
    return parser.parse_args()


def normalize_exact(value: Any) -> str:
    return unicodedata.normalize("NFKC", str(value or "")).strip()


def parse_jwt_source(token: str, default: str = "APP") -> str:
    try:
        parts = token.split(".")
        if len(parts) >= 2:
            payload_b64 = parts[1] + "=" * (-len(parts[1]) % 4)
            payload = json.loads(base64.urlsafe_b64decode(payload_b64))
            aud = payload.get("aud")
            if aud:
                return str(aud)
    except Exception:
        pass
    return default


def copy_column_widths(source_ws, target_ws) -> None:
    for column_letter, dimension in source_ws.column_dimensions.items():
        if dimension.width:
            target_ws.column_dimensions[column_letter].width = dimension.width


def classify_data_type(item: dict[str, Any]) -> str:
    raw_value = item.get("dataType")
    try:
        data_type = int(raw_value)
    except (TypeError, ValueError):
        data_type = None

    if data_type == 2:
        return "wps"
    if data_type == 3:
        return "circuit"
    if data_type is not None:
        return f"legacy_{data_type}"
    return "unknown"


def validate_token(client: httpx.Client, token: str) -> None:
    response = client.post(
        VALIDATE_URL,
        json={},
        headers={
            "app-token": token,
            "source": parse_jwt_source(token),
            "Content-Type": "application/json",
        },
    )
    response.raise_for_status()
    data = response.json()
    if not (
        isinstance(data, dict)
        and data.get("status") == 200
        and isinstance(data.get("data"), dict)
        and data["data"].get("userId")
    ):
        message = data.get("msg", "token 无效") if isinstance(data, dict) else "token 无效"
        raise RuntimeError(message)


def search_first(client: httpx.Client, *, token: str, query: str, page_size: int) -> dict[str, Any] | None:
    response = client.post(
        SEARCH_URL,
        json={
            "parentClassId": "0",
            "dataName": query,
            "pageNum": 1,
            "pageSize": page_size,
        },
        headers={
            "Content-Type": "application/json;charset=UTF-8",
            "app-token": token,
            "source": parse_jwt_source(token),
        },
    )
    response.raise_for_status()
    data = response.json()
    if data.get("status") != 200:
        raise RuntimeError(data.get("msg", "搜索接口返回异常"))
    items = data.get("data", {}).get("dataList", [])
    return items[0] if items else None


def build_audit_row(
    base_row: tuple[Any, ...],
    *,
    link_type: str,
    match_status: str,
    first_result_title: str = "",
    first_result_data_type: str = "",
) -> list[Any]:
    return list(base_row) + [link_type, match_status, first_result_title, first_result_data_type]


def add_sheet_with_rows(workbook: Workbook, *, name: str, header: list[Any], rows: list[list[Any]], source_ws) -> None:
    ws = workbook.create_sheet(title=name)
    ws.append(header)
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    copy_column_widths(source_ws, ws)


def add_summary_sheet(
    workbook: Workbook,
    *,
    candidate_size: int,
    target_size: int,
    exact_matches: int,
    counts: dict[str, int],
) -> None:
    ws = workbook.create_sheet(title="summary")
    ws.append(["metric", "value"])
    ws.append(["candidate_size", candidate_size])
    ws.append(["target_size", target_size])
    ws.append(["exact_match_count", exact_matches])
    for key in sorted(counts):
        ws.append([f"link_type:{key}", counts[key]])
    ws.freeze_panes = "A2"


def main() -> None:
    args = parse_args()
    token = os.getenv(args.token_env, "").strip()
    if not token:
        raise RuntimeError(f"请先设置环境变量 {args.token_env}")

    if args.candidate_size <= 0:
        raise ValueError("--candidate-size 必须大于 0")
    if args.target_size <= 0:
        raise ValueError("--target-size 必须大于 0")
    if args.target_size > args.candidate_size:
        raise ValueError("--target-size 不能大于 --candidate-size")

    input_path = Path(args.input).expanduser()
    output_path = Path(args.output).expanduser()
    if not input_path.exists():
        raise FileNotFoundError(f"主表不存在: {input_path}")

    source_wb = load_workbook(input_path, read_only=False, data_only=False)
    source_ws = source_wb[args.sheet] if args.sheet else source_wb.worksheets[0]
    rows = list(source_ws.iter_rows(values_only=True))
    if not rows:
        raise RuntimeError("主表为空")

    header = list(rows[0])
    data_rows = list(rows[1:])
    if len(data_rows) < args.candidate_size:
        raise RuntimeError(f"主表只有 {len(data_rows)} 条记录，不足以抽取 {args.candidate_size} 条")
    if "name" not in header:
        raise RuntimeError("主表缺少 name 列")
    name_index = header.index("name")

    rng = random.Random(args.seed)
    candidate_rows = rng.sample(data_rows, args.candidate_size)

    audit_header = header + ["link_type", "match_status", "first_result_title", "first_result_data_type"]
    audit_rows: list[list[Any]] = []
    target_rows: list[list[Any]] = []
    counts: dict[str, int] = {}

    timeout = httpx.Timeout(20.0, connect=10.0)
    with httpx.Client(timeout=timeout) as client:
        validate_token(client, token)
        for index, row in enumerate(candidate_rows, start=1):
            name = normalize_exact(row[name_index])
            link_type = "empty_name"
            match_status = "empty_name"
            first_result_title = ""
            first_result_data_type = ""

            if name:
                try:
                    first = search_first(client, token=token, query=name, page_size=args.page_size)
                    if first is None:
                        link_type = "no_results"
                        match_status = "no_results"
                    else:
                        first_result_title = normalize_exact(first.get("dataNameWs") or first.get("dataName"))
                        first_result_data_type = str(first.get("dataType") or "")
                        if first_result_title == name:
                            link_type = classify_data_type(first)
                            match_status = "exact_match"
                        else:
                            link_type = "not_exact_match"
                            match_status = "not_exact_match"
                except Exception as exc:
                    link_type = f"error: {exc}"
                    match_status = "error"

            counts[link_type] = counts.get(link_type, 0) + 1
            audit_rows.append(
                build_audit_row(
                    row,
                    link_type=link_type,
                    match_status=match_status,
                    first_result_title=first_result_title,
                    first_result_data_type=first_result_data_type,
                )
            )
            if match_status == "exact_match" and link_type in {"wps", "circuit"} and len(target_rows) < args.target_size:
                target_rows.append(list(row) + [link_type])

            print(
                f"[{index}/{args.candidate_size}] link_type={link_type} "
                f"match_status={match_status} target={len(target_rows)}/{args.target_size} "
                f"name={name[:60]}"
            )
            if index < args.candidate_size and args.delay > 0:
                time.sleep(args.delay)

    output_wb = Workbook()
    output_wb.remove(output_wb.active)
    add_sheet_with_rows(
        output_wb,
        name=f"target_{args.target_size}",
        header=header + ["link_type"],
        rows=target_rows,
        source_ws=source_ws,
    )
    add_sheet_with_rows(
        output_wb,
        name=f"audit_{args.candidate_size}",
        header=audit_header,
        rows=audit_rows,
        source_ws=source_ws,
    )
    add_summary_sheet(
        output_wb,
        candidate_size=args.candidate_size,
        target_size=args.target_size,
        exact_matches=len(target_rows),
        counts=counts,
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_wb.save(output_path)
    print(
        f"已输出: {output_path} "
        f"(target={len(target_rows)}/{args.target_size}, candidates={args.candidate_size})"
    )


if __name__ == "__main__":
    main()
