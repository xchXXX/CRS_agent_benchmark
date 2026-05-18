"""File-backed DocSearch benchmark runner and evaluator."""

from __future__ import annotations

import asyncio
import csv
import json
import mimetypes
import re
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from statistics import median
from typing import Any
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

REPO_ROOT = Path(__file__).resolve().parents[3]
BENCHMARK_ROOT = REPO_ROOT / "benchmarks" / "doc_search"
DATASETS_ROOT = BENCHMARK_ROOT / "datasets"
RUNS_ROOT = BENCHMARK_ROOT / "runs"


def utc_now_iso() -> str:
    return datetime.utcnow().replace(microsecond=0).isoformat() + "Z"


def normalize_doc_name(value: object) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"\.(txt|pdf|docx?|xlsx?|jpg|jpeg|png)$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"[\s_\-.,;:!?/\\()（）【】\[\]{}]+", "", text)
    return text


def doc_name_matches(candidate: object, gold: object) -> bool:
    candidate_norm = normalize_doc_name(candidate)
    gold_norm = normalize_doc_name(gold)
    if not candidate_norm or not gold_norm:
        return False
    return candidate_norm == gold_norm or gold_norm in candidate_norm or candidate_norm in gold_norm


def load_jsonl(path: Path) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    if not path.exists():
        return items
    with path.open("r", encoding="utf-8") as file:
        for line in file:
            line = line.strip()
            if line:
                items.append(json.loads(line))
    return items


def append_jsonl(path: Path, item: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as file:
        file.write(json.dumps(item, ensure_ascii=False) + "\n")


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def read_json(path: Path, default: dict[str, Any] | None = None) -> dict[str, Any]:
    if not path.exists():
        return default or {}
    return json.loads(path.read_text(encoding="utf-8"))


def _case_answerable(case: dict[str, Any]) -> bool | None:
    return (case.get("gold") or {}).get("answerable")


def _gold_names(case: dict[str, Any]) -> list[str]:
    gold = case.get("gold") or {}
    return [str(item) for item in gold.get("acceptable_doc_names") or [] if str(item or "").strip()]


def _input_query(case: dict[str, Any]) -> str:
    input_payload = case.get("input") or {}
    return str(input_payload.get("question_text") or "").strip()


def _input_images(case: dict[str, Any]) -> list[str]:
    input_payload = case.get("input") or {}
    return [str(item) for item in input_payload.get("image_paths") or [] if str(item or "").strip()]


def _image_evidence_summary_from_payloads(payloads: list[dict[str, Any]]) -> str:
    parts = []
    for payload in payloads:
        summary = str(payload.get("summary") or "").strip()
        if summary:
            parts.append(summary)
            continue
        visible_text = [str(item) for item in payload.get("visible_text") or [] if str(item or "").strip()]
        if visible_text:
            parts.append("图片文字识别：" + " / ".join(visible_text[:5]))
    return "\n".join(parts)


def _json_clone(value: Any) -> Any:
    if value is None:
        return None
    try:
        return json.loads(json.dumps(value, ensure_ascii=False, default=str))
    except Exception:
        return value


def _safe_model_dump(value: Any) -> dict[str, Any] | list[Any] | str | None:
    if value is None:
        return None
    if hasattr(value, "model_dump"):
        try:
            return _json_clone(value.model_dump(mode="json"))
        except Exception:
            try:
                return _json_clone(value.model_dump())
            except Exception:
                return str(value)
    if isinstance(value, (dict, list, tuple, str, int, float, bool)):
        return _json_clone(value)
    return str(value)


def _safe_case_context_snapshot(value: Any) -> dict[str, Any] | None:
    dumped = _safe_model_dump(value)
    return dumped if isinstance(dumped, dict) else None


def enrich_predictions_with_cases(
    cases: list[dict[str, Any]],
    predictions: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Backfill immutable case inputs for old runs that predate chain fields."""
    case_by_id = {str(case.get("case_id")): case for case in cases}
    enriched: list[dict[str, Any]] = []
    for prediction in predictions:
        item = dict(prediction)
        case = case_by_id.get(str(item.get("case_id"))) or {}
        if not item.get("case_snapshot") and case:
            item["case_snapshot"] = _json_clone(case)
        if not item.get("question_text"):
            item["question_text"] = _input_query(case)
        if not item.get("image_paths"):
            item["image_paths"] = _input_images(case)
        if not item.get("image_evidence_summary") and item.get("image_evidence"):
            item["image_evidence_summary"] = _image_evidence_summary_from_payloads(item.get("image_evidence") or [])
        runtime = dict(item.get("runtime") or {})
        search_snapshot = runtime.get("search_snapshot") if isinstance(runtime.get("search_snapshot"), dict) else {}
        if not item.get("effective_query"):
            item["effective_query"] = (
                runtime.get("effective_query")
                or search_snapshot.get("query")
                or search_snapshot.get("original_query")
                or item.get("question_text")
                or _input_query(case)
            )
        if not item.get("planned_queries"):
            item["planned_queries"] = list(runtime.get("planned_queries") or search_snapshot.get("planned_queries") or [])
        if item.get("image_paths") and not runtime.get("image_paths"):
            runtime["image_paths"] = item.get("image_paths")
        if item.get("planned_queries") and not runtime.get("planned_queries"):
            runtime["planned_queries"] = item.get("planned_queries")
        if item.get("image_evidence"):
            runtime["images_used_by_runner"] = runtime.get("images_used_by_runner", True)
        if runtime:
            item["runtime"] = runtime
        enriched.append(item)
    return enriched


def benchmark_case_status(case: dict[str, Any], prediction: dict[str, Any] | None) -> str:
    """Return the benchmark outcome using gold labels instead of model self-judgement."""
    if prediction and prediction.get("error"):
        return "执行错误"

    gold_answerable = _case_answerable(case)
    if gold_answerable is True:
        if prediction and prediction.get("best_rank_in_top_k") is not None:
            return "主榜命中"
        if prediction and prediction.get("best_rank_full") is not None:
            return "主榜外召回"
        return "未召回"

    if gold_answerable is False:
        results = []
        runtime: dict[str, Any] = {}
        prediction_answerable = None
        if prediction:
            results = prediction.get("results_scored") or prediction.get("results") or []
            runtime = prediction.get("runtime") or {}
            prediction_answerable = prediction.get("answerable")
        validity = runtime.get("validity") or {}
        is_correct_no_answer = (
            prediction_answerable is False
            or not results
            or validity.get("has_valid_results") is False
        )
        return "无资料正确" if is_correct_no_answer else "无资料误召回"

    if prediction and prediction.get("best_rank_in_top_k") is not None:
        return "主榜命中"
    if prediction and prediction.get("best_rank_full") is not None:
        return "主榜外召回"
    return "未召回"


def _result_doc_name(item: dict[str, Any]) -> str:
    return str(item.get("doc_name") or item.get("filename") or item.get("title") or item.get("path") or "").strip()


def _extract_results(envelope: dict[str, Any], *, track: str) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    data = envelope.get("data") if envelope.get("status") in {"ok", "need_clarify"} else envelope
    data = data or {}
    raw_results = list(data.get("results") or [])
    results = []
    for index, item in enumerate(raw_results, start=1):
        doc_name = _result_doc_name(item)
        if not doc_name:
            continue
        results.append(
            {
                "rank": index,
                "doc_name": doc_name,
                "doc_id": str(item.get("file_id") or item.get("doc_id") or "") or None,
                "score": item.get("score"),
                "path": item.get("physical_path") or item.get("hierarchy_full") or item.get("path"),
            }
        )
    runtime = {
        "search_method": data.get("search_method") or data.get("stats", {}).get("debug_info", {}).get("search_method"),
        "total": data.get("total") if track == "final_list" else len(results),
        "validity": data.get("validity"),
    }
    return results, runtime


def _extract_results_from_documents_content(content: dict[str, Any]) -> list[dict[str, Any]]:
    raw_results = list(content.get("results") or [])
    results = []
    for index, item in enumerate(raw_results, start=1):
        doc_name = _result_doc_name(item)
        if not doc_name:
            continue
        results.append(
            {
                "rank": index,
                "doc_name": doc_name,
                "doc_id": str(item.get("file_id") or item.get("doc_id") or "") or None,
                "score": item.get("score"),
                "path": item.get("physical_path") or item.get("hierarchy_full") or item.get("path"),
            }
        )
    return results


def _best_rank(results: list[dict[str, Any]], gold_names: list[str]) -> int | None:
    best: int | None = None
    for result in results:
        doc_name = result.get("doc_name")
        if any(doc_name_matches(doc_name, gold) for gold in gold_names):
            rank = int(result.get("rank") or 0)
            if rank > 0 and (best is None or rank < best):
                best = rank
    return best


def _matched_gold_details(results: list[dict[str, Any]], gold_names: list[str]) -> dict[str, Any]:
    matched_items: list[dict[str, Any]] = []
    matched_gold_names: list[str] = []
    matched_result_doc_names: list[str] = []
    seen_gold: set[str] = set()
    seen_result_doc: set[str] = set()

    for result in results:
        doc_name = str(result.get("doc_name") or "").strip()
        if not doc_name:
            continue
        matched_gold_for_result = [gold for gold in gold_names if doc_name_matches(doc_name, gold)]
        if not matched_gold_for_result:
            continue
        matched_items.append(
            {
                "rank": int(result.get("rank") or 0) or None,
                "doc_name": doc_name,
                "matched_gold_names": matched_gold_for_result,
            }
        )
        if doc_name not in seen_result_doc:
            seen_result_doc.add(doc_name)
            matched_result_doc_names.append(doc_name)
        for gold in matched_gold_for_result:
            if gold not in seen_gold:
                seen_gold.add(gold)
                matched_gold_names.append(gold)

    return {
        "matched_items": matched_items,
        "matched_gold_names": matched_gold_names,
        "matched_result_doc_names": matched_result_doc_names,
    }


def _build_prediction_payload(
    *,
    case_id: str,
    track: str,
    answerable: bool,
    all_results: list[dict[str, Any]],
    scored_results: list[dict[str, Any]],
    runtime: dict[str, Any],
    gold_names: list[str],
    error: str | None,
    question_text: str = "",
    image_paths: list[str] | None = None,
    image_evidence_summary: str = "",
    effective_query: str = "",
    planned_queries: list[dict[str, Any]] | None = None,
    trace_entries: list[dict[str, Any]] | None = None,
    image_evidence: list[dict[str, Any]] | None = None,
    case_snapshot: dict[str, Any] | None = None,
    image_inputs: list[dict[str, Any]] | None = None,
    request_payload: dict[str, Any] | None = None,
    response_payload: dict[str, Any] | None = None,
    search_snapshot: dict[str, Any] | None = None,
    case_context_before: dict[str, Any] | None = None,
    case_context_after: dict[str, Any] | None = None,
) -> dict[str, Any]:
    best_rank_full = _best_rank(all_results, gold_names)
    best_rank_in_top_k = _best_rank(scored_results, gold_names)
    matched_details = _matched_gold_details(all_results, gold_names)
    result = {
        "case_id": case_id,
        "track": track,
        "answerable": answerable,
        "question_text": question_text,
        "image_paths": image_paths or [],
        "image_evidence_summary": image_evidence_summary,
        "effective_query": effective_query or question_text,
        "planned_queries": planned_queries or [],
        "best_rank": best_rank_in_top_k,
        "best_rank_in_top_k": best_rank_in_top_k,
        "best_rank_full": best_rank_full,
        "hit_in_top_k": best_rank_in_top_k is not None,
        "results": scored_results,
        "results_scored": scored_results,
        "results_full": all_results,
        "returned_result_count": len(scored_results),
        "full_result_count": len(all_results),
        "matched_gold_names": matched_details["matched_gold_names"],
        "matched_result_doc_names": matched_details["matched_result_doc_names"],
        "matched_items": matched_details["matched_items"],
        "runtime": runtime,
        "error": error,
    }
    if trace_entries is not None:
        result["trace_entries"] = trace_entries
    if image_evidence is not None:
        result["image_evidence"] = image_evidence
    if case_snapshot is not None:
        result["case_snapshot"] = case_snapshot
    if image_inputs is not None:
        result["image_inputs"] = image_inputs
    if request_payload is not None:
        result["request_payload"] = request_payload
    if response_payload is not None:
        result["response_payload"] = response_payload
    if search_snapshot is not None:
        result["search_snapshot"] = search_snapshot
    if case_context_before is not None:
        result["case_context_before"] = case_context_before
    if case_context_after is not None:
        result["case_context_after"] = case_context_after
    return result


def evaluate_predictions(cases: list[dict[str, Any]], predictions: list[dict[str, Any]]) -> dict[str, Any]:
    prediction_by_case = {str(item.get("case_id")): item for item in predictions}
    answerable_cases = [case for case in cases if _case_answerable(case) is True]
    no_answer_cases = [case for case in cases if _case_answerable(case) is False]
    main_top_k_ranks: list[int | None] = []
    full_ranks: list[int | None] = []
    failures: list[dict[str, Any]] = []

    for case in answerable_cases:
        case_id = str(case.get("case_id"))
        prediction = prediction_by_case.get(case_id)
        gold_names = _gold_names(case)
        rank_in_top_k = prediction.get("best_rank_in_top_k") if prediction else None
        rank_full = prediction.get("best_rank_full") if prediction else None
        main_top_k_ranks.append(rank_in_top_k)
        full_ranks.append(rank_full)
        if rank_in_top_k is None:
            failures.append(
                {
                    "case_id": case_id,
                    "failure_type": "beyond_top_k" if rank_full is not None else "not_found_in_pool",
                    "question_text": _input_query(case),
                    "gold_doc_names": " | ".join(gold_names),
                    "best_rank": rank_full or "",
                    "best_rank_in_top_k": "",
                    "best_rank_full": rank_full or "",
                    "matched_gold_names": " | ".join((prediction or {}).get("matched_gold_names") or []),
                    "matched_result_doc_names": " | ".join((prediction or {}).get("matched_result_doc_names") or []),
                    "top_results": " | ".join(
                        result.get("doc_name", "") for result in (prediction or {}).get("results_scored", [])[:5]
                    ),
                }
            )

    def recall_at(k: int) -> float:
        if not full_ranks:
            return 0.0
        return sum(1 for rank in full_ranks if rank is not None and rank <= k) / len(full_ranks)

    reciprocal_ranks = [(1 / rank) if rank else 0 for rank in full_ranks]
    hit_ranks_main_top_k = [rank for rank in main_top_k_ranks if rank is not None]
    full_hit_ranks = [rank for rank in full_ranks if rank is not None]
    beyond_top_k_count = sum(
        1 for top_k_rank, full_rank in zip(main_top_k_ranks, full_ranks) if top_k_rank is None and full_rank is not None
    )
    not_found_in_pool_count = sum(1 for full_rank in full_ranks if full_rank is None)

    no_answer_correct = 0
    for case in no_answer_cases:
        prediction = prediction_by_case.get(str(case.get("case_id"))) or {}
        status_text = benchmark_case_status(case, prediction)
        results = prediction.get("results") or []
        if status_text == "无资料正确":
            no_answer_correct += 1
        elif prediction:
            failures.append(
                {
                    "case_id": str(case.get("case_id")),
                    "failure_type": "wrong_no_answer",
                    "question_text": _input_query(case),
                    "gold_doc_names": "",
                    "best_rank": "",
                    "best_rank_in_top_k": "",
                    "best_rank_full": "",
                    "matched_gold_names": "",
                    "matched_result_doc_names": "",
                    "top_results": " | ".join(result.get("doc_name", "") for result in results[:5]),
                }
            )

    by_task_type: dict[str, dict[str, Any]] = {}
    for case in answerable_cases:
        task_type = str((case.get("metadata") or {}).get("task_type") or "未分类")
        prediction = prediction_by_case.get(str(case.get("case_id"))) or {}
        rank = prediction.get("best_rank_full")
        bucket = by_task_type.setdefault(
            task_type,
            {
                "total": 0,
                "hit_at_5": 0,
                "hit_at_10": 0,
                "hit_at_50": 0,
                "hit_at_100": 0,
            },
        )
        bucket["total"] += 1
        bucket["hit_at_5"] += int(rank is not None and rank <= 5)
        bucket["hit_at_10"] += int(rank is not None and rank <= 10)
        bucket["hit_at_50"] += int(rank is not None and rank <= 50)
        bucket["hit_at_100"] += int(rank is not None and rank <= 100)
    for bucket in by_task_type.values():
        total = bucket["total"] or 1
        bucket["recall_at_5"] = bucket["hit_at_5"] / total
        bucket["recall_at_10"] = bucket["hit_at_10"] / total
        bucket["recall_at_50"] = bucket["hit_at_50"] / total
        bucket["recall_at_100"] = bucket["hit_at_100"] / total

    return {
        "summary": {
            "total_cases": len(cases),
            "answerable_cases": len(answerable_cases),
            "no_answer_cases": len(no_answer_cases),
            "evaluated_predictions": len(predictions),
            "recall_at_5": recall_at(5),
            "recall_at_10": recall_at(10),
            "recall_at_50": recall_at(50),
            "recall_at_100": recall_at(100),
            "mrr": sum(reciprocal_ranks) / len(reciprocal_ranks) if reciprocal_ranks else 0.0,
            "median_gold_rank": median(hit_ranks_main_top_k) if hit_ranks_main_top_k else None,
            "median_gold_rank_full": median(full_hit_ranks) if full_hit_ranks else None,
            "miss_rate": sum(1 for rank in full_ranks if rank is None) / len(full_ranks) if full_ranks else 0.0,
            "beyond_top_k_count": beyond_top_k_count,
            "beyond_top_k_rate": (beyond_top_k_count / len(main_top_k_ranks)) if main_top_k_ranks else 0.0,
            "not_found_in_pool_count": not_found_in_pool_count,
            "not_found_in_pool_rate": (not_found_in_pool_count / len(full_ranks)) if full_ranks else 0.0,
            "no_answer_accuracy": (no_answer_correct / len(no_answer_cases)) if no_answer_cases else None,
        },
        "by_task_type": by_task_type,
        "failures": failures,
    }


def write_failures_csv(path: Path, failures: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(
            file,
            fieldnames=[
                "case_id",
                "failure_type",
                "question_text",
                "gold_doc_names",
                "best_rank",
                "best_rank_in_top_k",
                "best_rank_full",
                "matched_gold_names",
                "matched_result_doc_names",
                "top_results",
            ],
        )
        writer.writeheader()
        writer.writerows(failures)


def write_markdown_report(
    path: Path,
    *,
    config: dict[str, Any],
    status: dict[str, Any],
    report: dict[str, Any],
) -> None:
    summary = report.get("summary") or {}
    by_task_type = report.get("by_task_type") or {}
    failures = report.get("failures") or []
    lines = [
        f"# Benchmark 测试报告: {config.get('run_id') or status.get('run_id') or '-'}",
        "",
        "## 基本信息",
        "",
        f"- 数据集: `{config.get('dataset_id') or '-'}`",
        f"- Track: `{config.get('track') or '-'}`",
        f"- 主榜 Top-K: `{config.get('top_k') or '-'}`",
        f"- 诊断候选池: `{config.get('diagnostic_pool_k') or '-'}`",
        f"- 创建时间: `{config.get('created_at') or status.get('created_at') or '-'}`",
        f"- 开始时间: `{status.get('started_at') or '-'}`",
        f"- 完成时间: `{status.get('finished_at') or '-'}`",
        f"- 状态: `{status.get('status') or '-'}`",
        "",
        "## 范围说明",
        "",
        f"- 主评测任务: `{((config.get('scope') or {}).get('primary')) or 'list_retrieval'}`",
        f"- 图片直接参与 runner: `{((config.get('scope') or {}).get('images_used_by_runner'))}`",
        f"- 澄清计入主分: `{((config.get('scope') or {}).get('clarification_in_main_score'))}`",
        f"- 完整排名候选池: `{((config.get('scope') or {}).get('full_rank_pool_k'))}`",
        "",
        "## 总体指标",
        "",
        f"- 总案例数: `{summary.get('total_cases', 0)}`",
        f"- 可答案例数: `{summary.get('answerable_cases', 0)}`",
        f"- 无资料案例数: `{summary.get('no_answer_cases', 0)}`",
        f"- Recall@5: `{summary.get('recall_at_5')}`",
        f"- Recall@10: `{summary.get('recall_at_10')}`",
        f"- Recall@50: `{summary.get('recall_at_50')}`",
        f"- Recall@100: `{summary.get('recall_at_100')}`",
        f"- MRR: `{summary.get('mrr')}`",
        f"- 主榜中位名次: `{summary.get('median_gold_rank')}`",
        f"- 诊断池中位名次: `{summary.get('median_gold_rank_full')}`",
        f"- Miss Rate: `{summary.get('miss_rate')}`",
        f"- 主榜外召回数: `{summary.get('beyond_top_k_count')}`",
        f"- 主榜外召回率: `{summary.get('beyond_top_k_rate')}`",
        f"- 诊断池未命中数: `{summary.get('not_found_in_pool_count')}`",
        f"- 诊断池未命中率: `{summary.get('not_found_in_pool_rate')}`",
        f"- No-answer Accuracy: `{summary.get('no_answer_accuracy')}`",
        "",
        "## 按问题类型",
        "",
    ]

    if by_task_type:
        lines.extend(
            [
                "| 问题类型 | 样本数 | Recall@5 | Recall@10 | Recall@50 | Recall@100 |",
                "| --- | ---: | ---: | ---: | ---: | ---: |",
            ]
        )
        for task_type, metrics in sorted(by_task_type.items(), key=lambda item: str(item[0])):
            lines.append(
                f"| {task_type} | {metrics.get('total', 0)} | "
                f"{metrics.get('recall_at_5', 0):.4f} | "
                f"{metrics.get('recall_at_10', 0):.4f} |"
                f"{metrics.get('recall_at_50', 0):.4f} | "
                f"{metrics.get('recall_at_100', 0):.4f} |"
            )
        lines.append("")
    else:
        lines.extend(["暂无分类指标。", ""])

    lines.extend(["## 失败样例摘要", ""])
    if failures:
        lines.extend(
            [
                "| Case ID | 类型 | 主榜名次 | 完整池名次 | 问题 | Gold | 命中资料 | Top Results |",
                "| --- | --- | ---: | ---: | --- | --- | --- | --- |",
            ]
        )
        for item in failures[:20]:
            question = str(item.get("question_text") or "").replace("\n", " ").replace("|", "/")
            gold = str(item.get("gold_doc_names") or "").replace("\n", " ").replace("|", "/")
            matched_docs = str(item.get("matched_result_doc_names") or "").replace("\n", " ").replace("|", "/")
            top_results = str(item.get("top_results") or "").replace("\n", " ").replace("|", "/")
            lines.append(
                f"| {item.get('case_id') or '-'} | {item.get('failure_type') or '-'} | "
                f"{item.get('best_rank_in_top_k') or '-'} | "
                f"{item.get('best_rank_full') or '-'} | "
                f"{question[:80]} | {gold[:80]} | {matched_docs[:80]} | {top_results[:120]} |"
            )
        if len(failures) > 20:
            lines.extend(["", f"其余 {len(failures) - 20} 条失败样例请查看 `failures.csv`。", ""])
        else:
            lines.append("")
    else:
        lines.extend(["没有失败样例。", ""])

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines), encoding="utf-8")


def write_excel_report(
    path: Path,
    *,
    config: dict[str, Any],
    status: dict[str, Any],
    report: dict[str, Any],
    cases: list[dict[str, Any]] | None = None,
    predictions: list[dict[str, Any]],
    events: list[dict[str, Any]] | None = None,
) -> None:
    workbook = Workbook()
    summary = report.get("summary") or {}
    events = events or []
    predictions = enrich_predictions_with_cases(cases or [], predictions) if cases else predictions
    case_by_id = {str(case.get("case_id")): case for case in (cases or [])}
    dark_fill = PatternFill(fill_type="solid", fgColor="122033")
    header_fill = PatternFill(fill_type="solid", fgColor="203044")
    section_fill = PatternFill(fill_type="solid", fgColor="EAF7F3")
    success_fill = PatternFill(fill_type="solid", fgColor="E8F7EF")
    warning_fill = PatternFill(fill_type="solid", fgColor="FFF4DE")
    danger_fill = PatternFill(fill_type="solid", fgColor="FDECEC")
    muted_fill = PatternFill(fill_type="solid", fgColor="F3F6FA")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(color="FFFFFF", bold=True, size=16)
    section_font = Font(color="0F172A", bold=True)
    muted_font = Font(color="64748B")
    wrap_alignment = Alignment(vertical="top", wrap_text=True)
    center_alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    thin_side = Side(style="thin", color="D7DEE8")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    def pct(value: Any) -> str:
        return f"{float(value) * 100:.1f}%" if isinstance(value, (int, float)) else "-"

    def fmt(value: Any) -> str:
        if value is None or value == "":
            return "-"
        if isinstance(value, bool):
            return "是" if value else "否"
        if isinstance(value, (dict, list)):
            return json.dumps(value, ensure_ascii=False)
        return str(value)

    def status_fill(value: str) -> PatternFill:
        if value in {"主榜命中", "无资料正确"}:
            return success_fill
        if value == "主榜外召回":
            return warning_fill
        if value in {"未召回", "无资料误召回", "执行错误"}:
            return danger_fill
        return muted_fill

    def style_header(row: int, sheet: Any, columns: list[str]) -> None:
        for col_index, _ in enumerate(columns, start=1):
            cell = sheet.cell(row=row, column=col_index)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border

    def style_table(sheet: Any, header_row: int = 1) -> None:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = wrap_alignment
                cell.border = border
        sheet.freeze_panes = f"A{header_row + 1}"
        sheet.auto_filter.ref = sheet.dimensions

    def add_title(sheet: Any, title: str, subtitle: str) -> None:
        sheet.merge_cells("A1:F1")
        sheet["A1"] = title
        sheet["A1"].fill = dark_fill
        sheet["A1"].font = title_font
        sheet["A1"].alignment = Alignment(vertical="center")
        sheet.row_dimensions[1].height = 30
        sheet.merge_cells("A2:F2")
        sheet["A2"] = subtitle
        sheet["A2"].fill = dark_fill
        sheet["A2"].font = Font(color="BFD8D0", size=11)
        sheet["A2"].alignment = Alignment(vertical="center")

    def append_section(sheet: Any, title: str) -> None:
        row = sheet.max_row + 2
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = sheet.cell(row=row, column=1, value=title)
        cell.fill = section_fill
        cell.font = section_font
        cell.alignment = Alignment(vertical="center")

    summary_sheet = workbook.active
    summary_sheet.title = "报告总览"
    add_title(
        summary_sheet,
        "资料搜索 Benchmark 测试报告",
        f"Run: {config.get('run_id') or status.get('run_id') or '-'}    Dataset: {config.get('dataset_id') or '-'}",
    )
    summary_sheet.append([])
    cards = [
        ("Recall@5", pct(summary.get("recall_at_5")), "目标资料进入前 5 的比例"),
        ("Recall@10", pct(summary.get("recall_at_10")), "目标资料进入前 10 的比例"),
        ("Recall@50", pct(summary.get("recall_at_50")), "目标资料进入前 50 的比例"),
        ("Recall@100", pct(summary.get("recall_at_100")), "目标资料进入前 100 的比例"),
        ("MRR", f"{summary.get('mrr', 0):.3f}" if isinstance(summary.get("mrr"), (int, float)) else "-", "越接近 1 排名越靠前"),
        ("诊断池未命中", str(summary.get("not_found_in_pool_count", 0)), "完整候选池也未找到目标"),
    ]
    for index, (label, value, note) in enumerate(cards, start=1):
        col = index
        label_cell = summary_sheet.cell(row=4, column=col, value=label)
        value_cell = summary_sheet.cell(row=5, column=col, value=value)
        note_cell = summary_sheet.cell(row=6, column=col, value=note)
        label_cell.fill = header_fill
        label_cell.font = header_font
        value_cell.fill = section_fill
        value_cell.font = Font(color="0F172A", bold=True, size=14)
        note_cell.fill = muted_fill
        note_cell.font = muted_font
        for cell in (label_cell, value_cell, note_cell):
            cell.alignment = center_alignment
            cell.border = border

    append_section(summary_sheet, "运行配置")
    config_start = summary_sheet.max_row + 1
    summary_sheet.append(["字段", "值", "说明"])
    style_header(config_start, summary_sheet, ["字段", "值", "说明"])
    for row in [
        ("状态", status.get("status") or "-", "completed 表示本次评测已结束"),
        ("Track", config.get("track") or "-", "production_flow 表示走真实资料搜索全链路"),
        ("主榜 Top-K", config.get("top_k") or "-", "仅用于主榜命中、主榜外召回和失败诊断。"),
        ("诊断候选池", config.get("diagnostic_pool_k") or "-", "固定评测指标 R@5/10/50/100 和完整池名次均基于该候选池。"),
        ("开始时间", status.get("started_at") or "-", ""),
        ("完成时间", status.get("finished_at") or "-", ""),
        ("样本数", summary.get("total_cases", 0), "包含可答案例和无资料案例"),
        ("可答案例", summary.get("answerable_cases", 0), "有人工标注正确资料的 case"),
        ("无资料案例", summary.get("no_answer_cases", 0), "人工标注不应返回资料的 case"),
    ]:
        summary_sheet.append(list(row))

    append_section(summary_sheet, "指标解释")
    metric_start = summary_sheet.max_row + 1
    summary_sheet.append(["指标", "当前值", "解释"])
    style_header(metric_start, summary_sheet, ["指标", "当前值", "解释"])
    for row in [
        ("Recall@K", f"R@5={pct(summary.get('recall_at_5'))} / R@50={pct(summary.get('recall_at_50'))}", "可答案例中，正确资料在诊断候选池中进入前 5/10/50/100 的比例。"),
        ("MRR", f"{summary.get('mrr', 0):.3f}" if isinstance(summary.get("mrr"), (int, float)) else "-", "正确资料排名倒数的平均值，越高表示越靠前。"),
        ("主榜中位名次", fmt(summary.get("median_gold_rank")), "只统计主榜 Top-K 命中的 case，观察主榜排序集中位置。"),
        ("诊断池中位名次", fmt(summary.get("median_gold_rank_full")), "在诊断候选池中找到目标资料时的中位排名。"),
        ("主榜外召回率", pct(summary.get("beyond_top_k_rate")), "能在诊断池找到，但没有进入主榜 Top-K，通常是 rerank 问题。"),
        ("诊断池未命中率", pct(summary.get("not_found_in_pool_rate")), "完整池也没有目标，通常是 query/索引/召回问题。"),
        ("No-answer Accuracy", pct(summary.get("no_answer_accuracy")), "无资料样例被系统正确判断为无有效资料的比例。"),
    ]:
        summary_sheet.append(list(row))

    for col_letter, width in {"A": 22, "B": 20, "C": 48, "D": 18, "E": 18, "F": 18}.items():
        summary_sheet.column_dimensions[col_letter].width = width
    summary_sheet.freeze_panes = "A8"

    task_sheet = workbook.create_sheet("按问题类型")
    task_columns = ["问题类型", "样本数", "Recall@5", "Recall@10", "Recall@50", "Recall@100"]
    task_sheet.append(task_columns)
    style_header(1, task_sheet, task_columns)
    for task_type, metrics in sorted((report.get("by_task_type") or {}).items(), key=lambda item: str(item[0])):
        task_sheet.append([
            task_type,
            metrics.get("total", 0),
            pct(metrics.get("recall_at_5", 0)),
            pct(metrics.get("recall_at_10", 0)),
            pct(metrics.get("recall_at_50", 0)),
            pct(metrics.get("recall_at_100", 0)),
        ])
    for cell in ("A", "B", "C", "D", "E", "F"):
        task_sheet.column_dimensions[cell].width = 20
    style_table(task_sheet)

    case_sheet = workbook.create_sheet("Case明细")
    case_columns = [
        "Case ID",
        "评测结论",
        "用户问题",
        "图片数",
        "图片识别摘要",
        "实际搜索问题",
        "评测链路",
        "是否应命中",
        "主榜名次",
        "完整池名次",
        "主榜命中",
        "主榜结果数",
        "完整池结果数",
        "命中的Gold资料",
        "命中的返回资料",
        "Top 5结果",
        "耗时(ms)",
        "响应类型",
        "诊断来源",
        "错误",
    ]
    case_sheet.append(case_columns)
    style_header(1, case_sheet, case_columns)
    for item in predictions:
        scored_results = item.get("results_scored") or item.get("results") or []
        case = case_by_id.get(str(item.get("case_id"))) or {}
        status_text = benchmark_case_status(case, item)
        image_count = len(item.get("image_paths") or []) or len(item.get("image_evidence") or [])
        case_sheet.append([
            item.get("case_id"),
            status_text,
            item.get("question_text"),
            image_count,
            item.get("image_evidence_summary"),
            item.get("effective_query"),
            item.get("track"),
            fmt(_case_answerable(case) if case else item.get("answerable")),
            item.get("best_rank_in_top_k") or "",
            item.get("best_rank_full") or "",
            fmt(item.get("hit_in_top_k")),
            item.get("returned_result_count"),
            item.get("full_result_count"),
            " | ".join(item.get("matched_gold_names") or []),
            " | ".join(item.get("matched_result_doc_names") or []),
            " | ".join(str(result.get("doc_name") or "") for result in scored_results[:5]),
            (item.get("runtime") or {}).get("latency_ms"),
            (item.get("runtime") or {}).get("response_type"),
            (item.get("runtime") or {}).get("diagnostic_rank_source"),
            item.get("error"),
        ])
        case_sheet.cell(row=case_sheet.max_row, column=2).fill = status_fill(status_text)
    for col_letter, width in {
        "A": 18,
        "B": 16,
        "C": 42,
        "D": 10,
        "E": 44,
        "F": 38,
        "G": 18,
        "H": 12,
        "I": 14,
        "J": 14,
        "K": 12,
        "L": 14,
        "M": 14,
        "N": 30,
        "O": 34,
        "P": 52,
        "Q": 12,
        "R": 16,
        "S": 18,
        "T": 30,
    }.items():
        case_sheet.column_dimensions[col_letter].width = width
    style_table(case_sheet)

    image_sheet = workbook.create_sheet("图片识别")
    image_columns = [
        "Case ID",
        "图片证据ID",
        "场景",
        "识别摘要",
        "车辆信息",
        "诊断信息",
        "可见文字",
        "建议查询",
        "置信度",
        "是否需确认",
    ]
    image_sheet.append(image_columns)
    style_header(1, image_sheet, image_columns)
    for item in predictions:
        for evidence in item.get("image_evidence") or []:
            vehicle = evidence.get("vehicle") if isinstance(evidence.get("vehicle"), dict) else {}
            diagnosis = evidence.get("diagnosis") if isinstance(evidence.get("diagnosis"), dict) else {}
            image_sheet.append([
                item.get("case_id"),
                evidence.get("image_evidence_id"),
                evidence.get("scene"),
                evidence.get("summary"),
                json.dumps(vehicle, ensure_ascii=False),
                json.dumps(diagnosis, ensure_ascii=False),
                " | ".join(evidence.get("visible_text") or []),
                " | ".join(evidence.get("suggested_queries") or []),
                evidence.get("confidence"),
                evidence.get("needs_user_confirm"),
            ])
    for col_letter, width in {"A": 18, "B": 24, "C": 18, "D": 48, "E": 38, "F": 38, "G": 42, "H": 42, "I": 12, "J": 18}.items():
        image_sheet.column_dimensions[col_letter].width = width
    style_table(image_sheet)

    plan_sheet = workbook.create_sheet("查询规划")
    plan_columns = ["Case ID", "查询序号", "规划查询", "置信度", "命中数", "实际搜索问题"]
    plan_sheet.append(plan_columns)
    style_header(1, plan_sheet, plan_columns)
    for item in predictions:
        planned_queries = item.get("planned_queries") or []
        if not planned_queries:
            plan_sheet.append([item.get("case_id"), "", "", "", "", item.get("effective_query")])
            continue
        for index, query in enumerate(planned_queries, start=1):
            query = query if isinstance(query, dict) else {"query": query}
            plan_sheet.append([
                item.get("case_id"),
                index,
                query.get("query"),
                query.get("confidence"),
                query.get("hit_count"),
                item.get("effective_query"),
            ])
    for col_letter, width in {"A": 18, "B": 12, "C": 48, "D": 14, "E": 12, "F": 48}.items():
        plan_sheet.column_dimensions[col_letter].width = width
    style_table(plan_sheet)

    input_sheet = workbook.create_sheet("案例输入")
    input_columns = [
        "Case ID",
        "用户问题",
        "Gold是否可答",
        "Gold资料",
        "图片路径",
        "图片输入元数据",
        "Case快照",
    ]
    input_sheet.append(input_columns)
    style_header(1, input_sheet, input_columns)
    for item in predictions:
        case_snapshot = item.get("case_snapshot") if isinstance(item.get("case_snapshot"), dict) else {}
        if not case_snapshot:
            case_snapshot = case_by_id.get(str(item.get("case_id"))) or {}
        gold_payload = case_snapshot.get("gold") if isinstance(case_snapshot.get("gold"), dict) else {}
        input_sheet.append([
            item.get("case_id"),
            item.get("question_text"),
            gold_payload.get("answerable"),
            " | ".join(str(doc) for doc in gold_payload.get("acceptable_doc_names") or []),
            " | ".join(item.get("image_paths") or []),
            json.dumps(item.get("image_inputs") or [], ensure_ascii=False),
            json.dumps(case_snapshot, ensure_ascii=False),
        ])
    for col_letter, width in {"A": 18, "B": 42, "C": 12, "D": 30, "E": 42, "F": 42, "G": 60}.items():
        input_sheet.column_dimensions[col_letter].width = width
    style_table(input_sheet)

    request_sheet = workbook.create_sheet("请求响应")
    request_columns = [
        "Case ID",
        "请求载荷",
        "响应类型",
        "业务",
        "响应载荷",
        "搜索快照",
        "Runtime",
    ]
    request_sheet.append(request_columns)
    style_header(1, request_sheet, request_columns)
    for item in predictions:
        runtime = item.get("runtime") if isinstance(item.get("runtime"), dict) else {}
        request_sheet.append([
            item.get("case_id"),
            json.dumps(item.get("request_payload") or {}, ensure_ascii=False),
            runtime.get("response_type"),
            runtime.get("business"),
            json.dumps(item.get("response_payload") or {}, ensure_ascii=False),
            json.dumps(item.get("search_snapshot") or {}, ensure_ascii=False),
            json.dumps(runtime, ensure_ascii=False),
        ])
    for col_letter, width in {"A": 18, "B": 48, "C": 14, "D": 14, "E": 60, "F": 60, "G": 60}.items():
        request_sheet.column_dimensions[col_letter].width = width
    style_table(request_sheet)

    context_sheet = workbook.create_sheet("上下文快照")
    context_columns = ["Case ID", "执行前Context", "执行后Context"]
    context_sheet.append(context_columns)
    style_header(1, context_sheet, context_columns)
    for item in predictions:
        context_sheet.append([
            item.get("case_id"),
            json.dumps(item.get("case_context_before") or {}, ensure_ascii=False),
            json.dumps(item.get("case_context_after") or {}, ensure_ascii=False),
        ])
    for col_letter, width in {"A": 18, "B": 72, "C": 72}.items():
        context_sheet.column_dimensions[col_letter].width = width
    style_table(context_sheet)

    trace_sheet = workbook.create_sheet("Trace明细")
    trace_columns = ["Case ID", "序号", "事件类型", "详情", "时间", "Payload"]
    trace_sheet.append(trace_columns)
    style_header(1, trace_sheet, trace_columns)
    for item in predictions:
        trace_entries = item.get("trace_entries") or []
        if not trace_entries:
            trace_sheet.append([item.get("case_id"), "", "", "", "", ""])
            continue
        for trace_entry in trace_entries:
            payload = trace_entry.get("payload") if isinstance(trace_entry, dict) else {}
            trace_sheet.append([
                item.get("case_id"),
                trace_entry.get("sequence_no"),
                trace_entry.get("event_type"),
                trace_entry.get("detail"),
                trace_entry.get("created_at"),
                json.dumps(payload, ensure_ascii=False),
            ])
    for col_letter, width in {"A": 18, "B": 10, "C": 24, "D": 32, "E": 24, "F": 72}.items():
        trace_sheet.column_dimensions[col_letter].width = width
    style_table(trace_sheet)

    full_results_sheet = workbook.create_sheet("完整候选")
    full_results_columns = ["Case ID", "排名", "资料名", "文档ID", "分数", "路径"]
    full_results_sheet.append(full_results_columns)
    style_header(1, full_results_sheet, full_results_columns)
    for item in predictions:
        full_results = item.get("results_full") or []
        if not full_results:
            full_results_sheet.append([item.get("case_id"), "", "", "", "", ""])
            continue
        for result in full_results:
            full_results_sheet.append([
                item.get("case_id"),
                result.get("rank"),
                result.get("doc_name"),
                result.get("doc_id"),
                result.get("score"),
                result.get("path"),
            ])
    for col_letter, width in {"A": 18, "B": 10, "C": 48, "D": 20, "E": 12, "F": 72}.items():
        full_results_sheet.column_dimensions[col_letter].width = width
    style_table(full_results_sheet)

    failures_sheet = workbook.create_sheet("失败样例")
    failure_columns = [
        "Case ID",
        "失败类型",
        "用户问题",
        "Gold资料",
        "主榜名次",
        "完整池名次",
        "命中的Gold资料",
        "命中的返回资料",
        "Top 结果",
    ]
    failures_sheet.append(failure_columns)
    style_header(1, failures_sheet, failure_columns)
    for item in report.get("failures") or []:
        failures_sheet.append([
            item.get("case_id"),
            item.get("failure_type"),
            item.get("question_text"),
            item.get("gold_doc_names"),
            item.get("best_rank_in_top_k"),
            item.get("best_rank_full"),
            item.get("matched_gold_names"),
            item.get("matched_result_doc_names"),
            item.get("top_results"),
        ])
    for col_letter, width in {"A": 18, "B": 18, "C": 42, "D": 28, "E": 16, "F": 16, "G": 24, "H": 32, "I": 48}.items():
        failures_sheet.column_dimensions[col_letter].width = width
    style_table(failures_sheet)

    events_sheet = workbook.create_sheet("执行日志")
    event_columns = ["时间", "事件类型", "消息", "Case ID", "事件载荷"]
    events_sheet.append(event_columns)
    style_header(1, events_sheet, event_columns)
    for item in events:
        payload = item.get("payload") if isinstance(item.get("payload"), dict) else {}
        events_sheet.append([
            item.get("ts"),
            item.get("event_type"),
            item.get("message"),
            payload.get("case_id"),
            json.dumps(payload, ensure_ascii=False) if payload else "",
        ])
    for col_letter, width in {"A": 22, "B": 24, "C": 42, "D": 18, "E": 72}.items():
        events_sheet.column_dimensions[col_letter].width = width
    style_table(events_sheet)

    glossary_sheet = workbook.create_sheet("字段说明")
    glossary_columns = ["字段", "位置", "说明"]
    glossary_sheet.append(glossary_columns)
    style_header(1, glossary_sheet, glossary_columns)
    for row in [
        ("评测结论", "Case明细", "主榜命中、主榜外召回、未召回、无资料正确、无资料误召回、执行错误。"),
        ("主榜名次", "Case明细", "正确资料在主榜 Top-K 候选中的最好排名，空值表示主榜未命中。"),
        ("完整池名次", "Case明细", "正确资料在诊断候选池中的最好排名，用于区分排序问题和召回问题。"),
        ("命中的Gold资料", "Case明细", "标注为正确资料且被返回结果匹配到的资料名，支持多个正确资料。"),
        ("实际搜索问题", "Case明细", "图片识别和 LLM 查询规划后，真正用于调用搜索接口的问题。"),
        ("图片识别摘要", "Case明细/图片识别", "图片证据分析的摘要，便于排查图片是否被正确理解。"),
        ("案例输入", "案例输入", "完整保留 case 输入、gold、图片路径、图片输入元数据和 case 快照。"),
        ("请求响应", "请求响应", "保留 benchmark 构造的 ChatRequest、ChatResponse、搜索快照和 runtime。"),
        ("上下文快照", "上下文快照", "执行前后 case context 对比，用于看图片识别和搜索结果是否写入上下文。"),
        ("Trace明细", "Trace明细", "逐条展开 case 内部 trace 事件，而不是只看全局执行日志。"),
        ("完整候选", "完整候选", "展开完整诊断候选池，用于判断是主榜裁切问题还是底层召回问题。"),
        ("Recall@K", "报告总览/按问题类型", "可答案例中，正确资料在诊断候选池中进入前 5/10/50/100 的比例。"),
        ("MRR", "报告总览", "正确资料排名倒数的平均值，越高越好。"),
        ("主榜外召回", "报告总览/失败样例", "诊断池能找到目标，但排序没有进入主榜 Top-K。"),
        ("诊断池未命中", "报告总览/失败样例", "扩大到诊断候选池后仍找不到目标，优先排查召回、索引或 query。"),
    ]:
        glossary_sheet.append(list(row))
    for col_letter, width in {"A": 22, "B": 20, "C": 72}.items():
        glossary_sheet.column_dimensions[col_letter].width = width
    style_table(glossary_sheet)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


@dataclass
class DatasetInfo:
    dataset_id: str
    path: Path
    case_count: int
    answerable_count: int
    no_answer_count: int
    card: str
    updated_at: str | None

    def to_dict(self) -> dict[str, Any]:
        return {
            "dataset_id": self.dataset_id,
            "case_count": self.case_count,
            "answerable_count": self.answerable_count,
            "no_answer_count": self.no_answer_count,
            "path": str(self.path),
            "card": self.card,
            "updated_at": self.updated_at,
        }


class DocSearchBenchmarkStore:
    def __init__(self, *, datasets_root: Path = DATASETS_ROOT, runs_root: Path = RUNS_ROOT):
        self.datasets_root = datasets_root
        self.runs_root = runs_root
        self.datasets_root.mkdir(parents=True, exist_ok=True)
        self.runs_root.mkdir(parents=True, exist_ok=True)

    def list_datasets(self) -> list[dict[str, Any]]:
        datasets: list[DatasetInfo] = []
        for dataset_dir in sorted(self.datasets_root.iterdir() if self.datasets_root.exists() else []):
            if not dataset_dir.is_dir():
                continue
            cases_path = dataset_dir / "cases.jsonl"
            if not cases_path.exists():
                continue
            cases = load_jsonl(cases_path)
            stat = cases_path.stat()
            datasets.append(
                DatasetInfo(
                    dataset_id=dataset_dir.name,
                    path=cases_path,
                    case_count=len(cases),
                    answerable_count=sum(1 for case in cases if _case_answerable(case) is True),
                    no_answer_count=sum(1 for case in cases if _case_answerable(case) is False),
                    card=(dataset_dir / "dataset_card.md").read_text(encoding="utf-8")
                    if (dataset_dir / "dataset_card.md").exists()
                    else "",
                    updated_at=datetime.fromtimestamp(stat.st_mtime).isoformat(),
                )
            )
        return [item.to_dict() for item in datasets]

    def load_cases(self, dataset_id: str) -> list[dict[str, Any]]:
        dataset_path = self.dataset_path(dataset_id)
        if not dataset_path.exists():
            raise FileNotFoundError(f"Dataset not found: {dataset_id}")
        return load_jsonl(dataset_path)

    def dataset_path(self, dataset_id: str) -> Path:
        safe_id = Path(dataset_id).name
        return self.datasets_root / safe_id / "cases.jsonl"

    def create_run(self, *, dataset_id: str, track: str, top_k: int, created_by: str | None = None) -> str:
        run_id = f"run_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}_{uuid4().hex[:8]}"
        run_dir = self.run_dir(run_id)
        run_dir.mkdir(parents=True, exist_ok=False)
        diagnostic_pool_k = max(top_k, 100)
        config = {
            "run_id": run_id,
            "dataset_id": dataset_id,
            "track": track,
            "top_k": top_k,
            "diagnostic_pool_k": diagnostic_pool_k,
            "created_by": created_by,
            "created_at": utc_now_iso(),
            "scope": {
                "primary": "list_retrieval",
                "images_used_by_runner": track == "production_flow",
                "clarification_in_main_score": False,
                "full_rank_pool_k": diagnostic_pool_k,
            },
        }
        write_json(run_dir / "config.json", config)
        write_json(
            run_dir / "status.json",
            {
                **config,
                "status": "queued",
                "progress": {"total": 0, "completed": 0, "failed": 0},
                "started_at": None,
                "finished_at": None,
                "summary": None,
                "error": None,
            },
        )
        return run_id

    def run_dir(self, run_id: str) -> Path:
        return self.runs_root / Path(run_id).name

    def list_runs(self) -> list[dict[str, Any]]:
        runs = []
        for run_dir in sorted(self.runs_root.iterdir() if self.runs_root.exists() else [], reverse=True):
            if not run_dir.is_dir():
                continue
            status = read_json(run_dir / "status.json", {})
            if status:
                runs.append(status)
        return runs

    def get_run_detail(self, run_id: str) -> dict[str, Any]:
        run_dir = self.run_dir(run_id)
        status = read_json(run_dir / "status.json", {})
        if not status:
            raise FileNotFoundError(f"Run not found: {run_id}")
        config = read_json(run_dir / "config.json", {})
        predictions = load_jsonl(run_dir / "predictions.jsonl")[-200:]
        dataset_id = str(config.get("dataset_id") or "")
        if dataset_id:
            try:
                predictions = enrich_predictions_with_cases(self.load_cases(dataset_id), predictions)
            except FileNotFoundError:
                pass
        return {
            "status": status,
            "config": config,
            "report": read_json(run_dir / "report.json", {}),
            "events": load_jsonl(run_dir / "events.jsonl")[-300:],
            "predictions": predictions,
        }

    def update_status(self, run_id: str, patch: dict[str, Any]) -> dict[str, Any]:
        path = self.run_dir(run_id) / "status.json"
        status = read_json(path, {})
        status.update(patch)
        write_json(path, status)
        return status

    def log_event(self, run_id: str, event_type: str, message: str, **payload: Any) -> None:
        append_jsonl(
            self.run_dir(run_id) / "events.jsonl",
            {
                "ts": utc_now_iso(),
                "event_type": event_type,
                "message": message,
                "payload": payload,
            },
        )


class DocSearchBenchmarkRunner:
    def __init__(self, *, store: DocSearchBenchmarkStore, runtime_deps: Any):
        self.store = store
        self.runtime_deps = runtime_deps

    async def run(self, run_id: str, *, resume: bool = False) -> None:
        config = read_json(self.store.run_dir(run_id) / "config.json", {})
        dataset_id = str(config.get("dataset_id") or "")
        track = str(config.get("track") or "production_flow")
        top_k = int(config.get("top_k") or 20)
        diagnostic_pool_k = int(config.get("diagnostic_pool_k") or max(top_k, 100))
        all_cases = self.store.load_cases(dataset_id)
        predictions_path = self.store.run_dir(run_id) / "predictions.jsonl"

        cases = all_cases
        completed = 0
        failed = 0
        if resume:
            existing_predictions = load_jsonl(predictions_path)
            processed_case_ids = {str(p.get("case_id")) for p in existing_predictions}
            cases = [c for c in all_cases if str(c.get("case_id")) not in processed_case_ids]
            completed = len(processed_case_ids)
            failed = sum(1 for p in existing_predictions if p.get("error"))
            self.store.log_event(
                run_id,
                "run_resumed",
                f"Benchmark run resumed, skipping {len(processed_case_ids)} already processed cases",
                total_cases=len(all_cases),
                remaining_cases=len(cases),
                completed_before=completed,
                failed_before=failed,
            )

        self.store.update_status(
            run_id,
            {
                "status": "running",
                "started_at": utc_now_iso() if not resume else read_json(self.store.run_dir(run_id) / "status.json", {}).get("started_at"),
                "progress": {"total": len(all_cases), "completed": completed, "failed": failed},
            },
        )
        if not resume:
            self.store.log_event(
                run_id,
                "run_started",
                "Benchmark run started",
                dataset_id=dataset_id,
                track=track,
                top_k=top_k,
                diagnostic_pool_k=diagnostic_pool_k,
            )

        try:
            from app.agent.adapters.legacy_doc_search_adapter import LegacyDocSearchAdapter

            adapter = LegacyDocSearchAdapter(self.runtime_deps)
            for case in cases:
                case_id = str(case.get("case_id"))
                self.store.log_event(
                    run_id,
                    "case_started",
                    f"{case_id} started",
                    case_id=case_id,
                    question_text=_input_query(case),
                    image_count=len(_input_images(case)),
                    gold_answerable=_case_answerable(case),
                    gold_doc_names=_gold_names(case),
                )
                prediction = await self._run_case(
                    adapter=adapter,
                    case=case,
                    track=track,
                    top_k=top_k,
                    diagnostic_pool_k=diagnostic_pool_k,
                )
                trace_entries = prediction.get("trace_entries", []) or []
                append_jsonl(predictions_path, prediction)
                for trace_entry in trace_entries:
                    self.store.log_event(
                        run_id,
                        f"trace:{trace_entry.get('event_type')}",
                        trace_entry.get("detail") or trace_entry.get("event_type") or "trace",
                        case_id=case.get("case_id"),
                        trace=trace_entry,
                    )
                completed += 1
                if prediction.get("error"):
                    failed += 1
                self.store.update_status(
                    run_id,
                    {"progress": {"total": len(all_cases), "completed": completed, "failed": failed}},
                )
                self.store.log_event(
                    run_id,
                    "case_completed",
                    f"{case_id} completed",
                    case_id=case_id,
                    best_rank=prediction.get("best_rank"),
                    best_rank_in_top_k=prediction.get("best_rank_in_top_k"),
                    best_rank_full=prediction.get("best_rank_full"),
                    result_count=len(prediction.get("results") or []),
                    full_result_count=len(prediction.get("results_full") or []),
                    response_type=((prediction.get("runtime") or {}).get("response_type")),
                    business=((prediction.get("runtime") or {}).get("business")),
                    effective_query=prediction.get("effective_query"),
                    error=prediction.get("error"),
                )

            predictions = load_jsonl(predictions_path)
            report = evaluate_predictions(all_cases, predictions)
            self._write_run_reports(run_id, config=config, all_cases=all_cases, predictions=predictions, report=report)
            self.store.update_status(
                run_id,
                {
                    "status": "completed",
                    "finished_at": utc_now_iso(),
                    "summary": report.get("summary"),
                },
            )
            self.store.log_event(run_id, "run_completed", "Benchmark run completed", summary=report.get("summary"))
        except asyncio.CancelledError:
            # Graceful pause: save what we have so far
            predictions = load_jsonl(predictions_path)
            if predictions:
                report = evaluate_predictions(all_cases, predictions)
                self._write_run_reports(run_id, config=config, all_cases=all_cases, predictions=predictions, report=report)
            self.store.update_status(
                run_id,
                {
                    "status": "paused",
                    "finished_at": utc_now_iso(),
                    "summary": report.get("summary") if predictions else None,
                },
            )
            self.store.log_event(run_id, "run_paused", "Benchmark run paused by user", completed=completed, failed=failed)
            raise
        except Exception as exc:
            self.store.update_status(
                run_id,
                {
                    "status": "failed",
                    "finished_at": utc_now_iso(),
                    "error": str(exc),
                },
            )
            self.store.log_event(run_id, "run_failed", "Benchmark run failed", error=str(exc))

    def _write_run_reports(
        self,
        run_id: str,
        *,
        config: dict[str, Any],
        all_cases: list[dict[str, Any]],
        predictions: list[dict[str, Any]],
        report: dict[str, Any],
    ) -> None:
        write_json(self.store.run_dir(run_id) / "report.json", report)
        write_failures_csv(self.store.run_dir(run_id) / "failures.csv", report.get("failures") or [])
        write_markdown_report(
            self.store.run_dir(run_id) / "report.md",
            config=config,
            status=read_json(self.store.run_dir(run_id) / "status.json", {}),
            report=report,
        )
        write_excel_report(
            self.store.run_dir(run_id) / "report.xlsx",
            config=config,
            status=read_json(self.store.run_dir(run_id) / "status.json", {}),
            report=report,
            cases=all_cases,
            predictions=predictions,
            events=load_jsonl(self.store.run_dir(run_id) / "events.jsonl"),
        )

    async def _run_case(
        self,
        *,
        adapter: Any,
        case: dict[str, Any],
        track: str,
        top_k: int,
        diagnostic_pool_k: int,
    ) -> dict[str, Any]:
        case_id = str(case.get("case_id"))
        query = _input_query(case)
        image_paths = _input_images(case)
        started = time.perf_counter()
        if track == "production_flow":
            return await self._run_case_via_production_flow(
                adapter=adapter,
                case=case,
                top_k=top_k,
                diagnostic_pool_k=diagnostic_pool_k,
                started=started,
            )

        error = None
        envelope: dict[str, Any] = {}
        try:
            if track == "raw_retrieval":
                envelope = await adapter.search_raw(query=query, top_k=diagnostic_pool_k)
            else:
                envelope = await adapter.search(query=query, top_k=diagnostic_pool_k)
        except Exception as exc:
            error = str(exc)
            envelope = {"status": "failed", "data": {"message": str(exc)}}

        latency_ms = round((time.perf_counter() - started) * 1000, 2)
        all_results, runtime = _extract_results(envelope, track=track)
        scored_results = all_results[:top_k]
        runtime["latency_ms"] = latency_ms
        runtime["image_paths"] = image_paths
        runtime["images_used_by_runner"] = False
        runtime["diagnostic_pool_k"] = diagnostic_pool_k
        gold_names = _gold_names(case)

        validity = runtime.get("validity") or {}
        answerable = bool(scored_results)
        if validity.get("has_valid_results") is False:
            answerable = False

        return _build_prediction_payload(
            case_id=case_id,
            track=track,
            answerable=answerable,
            all_results=all_results,
            scored_results=scored_results,
            runtime=runtime,
            gold_names=gold_names,
            error=error or (envelope.get("data") or {}).get("message") if envelope.get("status") == "failed" else None,
            question_text=query,
            image_paths=image_paths,
            effective_query=query,
            planned_queries=[{"query": query, "confidence": 1.0}],
        )

    async def _run_case_via_production_flow(
        self,
        *,
        adapter: Any,
        case: dict[str, Any],
        top_k: int,
        diagnostic_pool_k: int,
        started: float,
    ) -> dict[str, Any]:
        from app.agent.adapters.doc_search_response_adapter import DocSearchResponseAdapter
        from app.agent.domain.image_evidence import (
            ImageEvidenceImageInput,
            ImageEvidenceRequest,
            ImageEvidenceService,
        )
        from app.agent.runtime.service import AgentLoopService
        from app.schemas.chat import ChatRequest

        case_id = str(case.get("case_id"))
        query = _input_query(case)
        image_paths = _input_images(case)
        runtime_deps = self.runtime_deps.clone_for_request(tracer=self.runtime_deps.tracer.fork())
        service = AgentLoopService(deps=runtime_deps)
        session_id = f"benchmark_{case_id}_{uuid4().hex[:8]}"
        evidence_payloads: list[dict[str, Any]] = []
        image_inputs_meta: list[dict[str, Any]] = []
        image_evidence_summary = ""
        error: str | None = None
        request_payload: dict[str, Any] | None = None
        response_payload: dict[str, Any] | None = None
        search_snapshot: dict[str, Any] | None = None
        case_context_before = _safe_case_context_snapshot(getattr(runtime_deps, "case_context", None))
        case_context_after: dict[str, Any] | None = None

        try:
            if image_paths:
                image_inputs = []
                for image_path in image_paths:
                    path = Path(image_path)
                    content_type, _ = mimetypes.guess_type(str(path))
                    image_size_bytes = path.stat().st_size if path.exists() else None
                    image_inputs_meta.append(
                        {
                            "path": str(path),
                            "filename": path.name,
                            "content_type": content_type or "image/jpeg",
                            "size_bytes": image_size_bytes,
                        }
                    )
                    image_inputs.append(
                        ImageEvidenceImageInput(
                            filename=path.name,
                            content=path.read_bytes(),
                            content_type=content_type or "image/jpeg",
                        )
                    )
                runtime_deps.tracer.trace(
                    event_type="benchmark_image_evidence_started",
                    session_id=session_id,
                    detail="Benchmark image evidence analysis started",
                    payload={
                        "case_id": case_id,
                        "image_count": len(image_inputs),
                        "filenames": [image.filename for image in image_inputs],
                    },
                )
                evidence_result = await ImageEvidenceService(config_service=runtime_deps.config_service).analyze(
                    ImageEvidenceRequest(images=image_inputs, user_prompt=query)
                )
                if not evidence_result.success or evidence_result.evidence is None:
                    error = (
                        evidence_result.error.get("message")
                        if isinstance(evidence_result.error, dict)
                        else "图片证据分析失败"
                    )
                    runtime_deps.tracer.trace(
                        event_type="benchmark_image_evidence_failed",
                        session_id=session_id,
                        detail="Benchmark image evidence analysis failed",
                        payload={
                            "case_id": case_id,
                            "image_count": len(image_inputs),
                            "error": error,
                        },
                    )
                    raise RuntimeError(error)
                evidence_payloads.append(evidence_result.evidence.model_dump(mode="json"))
                image_evidence_summary = AgentLoopService._build_image_evidence_summary(evidence_payloads)
                runtime_deps.tracer.trace(
                    event_type="benchmark_image_evidence_succeeded",
                    session_id=session_id,
                    detail="Benchmark image evidence analysis succeeded",
                    payload={
                        "case_id": case_id,
                        "image_count": len(image_inputs),
                        "evidence_count": len(evidence_payloads),
                        "summary": image_evidence_summary,
                        "confidence": evidence_payloads[0].get("confidence") if evidence_payloads else None,
                        "scene": evidence_payloads[0].get("scene") if evidence_payloads else None,
                        "suggested_queries": evidence_payloads[0].get("suggested_queries") if evidence_payloads else [],
                    },
                )

            request = ChatRequest(
                message=query,
                session_id=session_id,
                mode="doc_search",
                client_type="benchmark",
                context={"image_evidences": evidence_payloads} if evidence_payloads else {},
            )
            request_payload = _safe_model_dump(request)
            response = await service.process(request, runtime_deps=runtime_deps)
            response_payload = _safe_model_dump(response)
            case_context_after = _safe_case_context_snapshot(getattr(runtime_deps, "case_context", None))
        except Exception as exc:
            latency_ms = round((time.perf_counter() - started) * 1000, 2)
            return _build_prediction_payload(
                case_id=case_id,
                track="production_flow",
                answerable=False,
                all_results=[],
                scored_results=[],
                runtime={
                    "latency_ms": latency_ms,
                    "image_paths": image_paths,
                    "images_used_by_runner": True,
                    "response_type": "error",
                },
                gold_names=_gold_names(case),
                error=str(exc),
                question_text=query,
                image_paths=image_paths,
                image_evidence_summary=image_evidence_summary,
                effective_query=query,
                planned_queries=[],
                image_evidence=evidence_payloads,
                case_snapshot=_json_clone(case),
                image_inputs=image_inputs_meta,
                request_payload=request_payload,
                response_payload=response_payload,
                search_snapshot=search_snapshot,
                case_context_before=case_context_before,
                case_context_after=_safe_case_context_snapshot(getattr(runtime_deps, "case_context", None)),
            )

        latency_ms = round((time.perf_counter() - started) * 1000, 2)
        all_results: list[dict[str, Any]] = []
        diagnostic_snapshot: dict[str, Any] | None = None
        runtime: dict[str, Any] = {
            "latency_ms": latency_ms,
            "image_paths": image_paths,
            "images_used_by_runner": True,
            "response_type": response.type,
            "business": response.business,
            "response_metadata": response.metadata,
            "need_clarify": response.need_clarify,
        }

        if response.type == "documents" and isinstance(response.content, dict):
            all_results = _extract_results_from_documents_content(response.content)
            runtime["planned_queries"] = response.content.get("planned_queries") or []
            runtime["query_plan_rationale"] = response.content.get("query_plan_rationale") or ""
            runtime["effective_query"] = response.content.get("query") or query
            runtime["result_count_before_top_k"] = len(all_results)
            search_snapshot = _json_clone(response.content)
            answerable = bool(all_results[:top_k])
        elif response.type == "ask_user":
            tool_call_id = (
                getattr(response.ask_user, "tool_call_id", None)
                or (response.metadata or {}).get("tool_call_id")
            )
            deferred_state = (
                runtime_deps.deferred_state_store.load(session_id=session_id, tool_call_id=tool_call_id)
                if tool_call_id
                else None
            )
            snapshot = DocSearchResponseAdapter.resolve_search_snapshot(deferred_state) if deferred_state else None
            if isinstance(snapshot, dict):
                diagnostic_snapshot = snapshot
                search_snapshot = _json_clone(snapshot)
                all_results = _extract_results_from_documents_content(
                    DocSearchResponseAdapter.build_documents_content(snapshot)
                )
                runtime["search_snapshot"] = {
                    "planned_queries": snapshot.get("planned_queries") or [],
                    "query_plan_rationale": snapshot.get("query_plan_rationale") or "",
                    "result_count": len(snapshot.get("results") or []),
                }
                runtime["effective_query"] = snapshot.get("query") or snapshot.get("original_query") or query
            runtime["ask_user"] = response.ask_user.model_dump(mode="json") if response.ask_user is not None else {}
            runtime["result_count_before_top_k"] = len(all_results)
            answerable = bool(all_results[:top_k])
        else:
            answerable = False

        trace_entries = runtime_deps.tracer.entries()
        runtime["trace_event_count"] = len(trace_entries)
        runtime["diagnostic_pool_k"] = diagnostic_pool_k
        runtime["result_count_before_top_k"] = len(all_results)
        runtime["image_input_count"] = len(image_inputs_meta)
        gold_names = _gold_names(case)
        scored_results = all_results[:top_k]
        diagnostic_results = all_results

        if diagnostic_pool_k > len(all_results):
            diagnostic_results = await self._expand_production_flow_diagnostic_results(
                adapter=adapter,
                query=query,
                response=response,
                diagnostic_snapshot=diagnostic_snapshot,
                current_results=all_results,
                diagnostic_pool_k=diagnostic_pool_k,
            )
            runtime["diagnostic_result_count"] = len(diagnostic_results)
            runtime["diagnostic_rank_source"] = "expanded_pool"
        else:
            runtime["diagnostic_result_count"] = len(diagnostic_results)
            runtime["diagnostic_rank_source"] = "returned_results"

        return _build_prediction_payload(
            case_id=case_id,
            track="production_flow",
            answerable=answerable,
            all_results=diagnostic_results,
            scored_results=scored_results,
            runtime=runtime,
            gold_names=gold_names,
            error=error,
            question_text=query,
            image_paths=image_paths,
            image_evidence_summary=image_evidence_summary,
            effective_query=str(runtime.get("effective_query") or query),
            planned_queries=list(runtime.get("planned_queries") or (runtime.get("search_snapshot") or {}).get("planned_queries") or []),
            trace_entries=[
                {
                    "sequence_no": entry.sequence_no,
                    "event_type": entry.event_type,
                    "detail": entry.detail,
                    "payload": entry.payload,
                    "created_at": entry.created_at,
                }
                for entry in trace_entries
            ],
            image_evidence=evidence_payloads,
            case_snapshot=_json_clone(case),
            image_inputs=image_inputs_meta,
            request_payload=request_payload,
            response_payload=response_payload,
            search_snapshot=search_snapshot,
            case_context_before=case_context_before,
            case_context_after=case_context_after,
        )

    async def _expand_production_flow_diagnostic_results(
        self,
        *,
        adapter: Any,
        query: str,
        response: Any,
        diagnostic_snapshot: dict[str, Any] | None,
        current_results: list[dict[str, Any]],
        diagnostic_pool_k: int,
    ) -> list[dict[str, Any]]:
        from app.agent.runtime.service import AgentLoopService, DocSearchExecutedQuery

        diagnostic_query = query
        planned_queries: list[dict[str, Any]] = []
        rationale = ""

        if response.type == "documents" and isinstance(response.content, dict):
            planned_queries = list(response.content.get("planned_queries") or [])
            rationale = str(response.content.get("query_plan_rationale") or "").strip()
            diagnostic_query = str(response.content.get("query") or diagnostic_query).strip() or diagnostic_query
        elif response.type == "ask_user":
            planned_queries = list((diagnostic_snapshot or {}).get("planned_queries") or [])
            rationale = str((diagnostic_snapshot or {}).get("query_plan_rationale") or "").strip()
            diagnostic_query = str(
                (diagnostic_snapshot or {}).get("original_query")
                or (diagnostic_snapshot or {}).get("query")
                or diagnostic_query
            ).strip() or diagnostic_query

        executed_queries: list[DocSearchExecutedQuery] = []
        if planned_queries:
            for item in planned_queries:
                planned_query = str(item.get("query") or "").strip()
                if not planned_query:
                    continue
                executed_queries.append(
                    DocSearchExecutedQuery(
                        query=planned_query,
                        confidence=float(item.get("confidence") or 0.5),
                    )
                )
            if executed_queries:
                diagnostic_query = executed_queries[0].query

        try:
            if executed_queries:
                envelopes = []
                for query_info in executed_queries:
                    envelope = await adapter.search_raw(query=query_info.query, top_k=diagnostic_pool_k)
                    envelopes.append((query_info, envelope))
                merged_envelope = AgentLoopService._merge_doc_search_envelopes(
                    envelopes,
                    primary_query=diagnostic_query,
                    rationale=rationale,
                )
                expanded_results, _ = _extract_results(merged_envelope, track="raw_retrieval")
            else:
                envelope = await adapter.search_raw(query=diagnostic_query, top_k=diagnostic_pool_k)
                expanded_results, _ = _extract_results(envelope, track="raw_retrieval")
            if expanded_results:
                return expanded_results
        except Exception:
            return current_results

        return current_results


def schedule_benchmark_run(*, store: DocSearchBenchmarkStore, runtime_deps: Any, run_id: str, resume: bool = False) -> asyncio.Task:
    runner = DocSearchBenchmarkRunner(store=store, runtime_deps=runtime_deps)
    return asyncio.create_task(runner.run(run_id, resume=resume))


async def pause_benchmark_run(*, task_map: dict[str, asyncio.Task], run_id: str) -> dict[str, Any]:
    task = task_map.get(run_id)
    if task is None:
        return {"run_id": run_id, "status": "not_found", "message": "运行任务不存在"}
    if task.done():
        return {"run_id": run_id, "status": "already_done", "message": "运行任务已结束"}
    task.cancel()
    # Wait a short moment for the cancellation to propagate and status to be written
    try:
        await asyncio.wait_for(task, timeout=5.0)
    except (asyncio.CancelledError, asyncio.TimeoutError):
        pass
    return {"run_id": run_id, "status": "pausing", "message": "正在暂停运行"}
