"""Excel-backed repair knowledge catalog service."""

from __future__ import annotations

from dataclasses import dataclass
from functools import cached_property
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook


QUESTION_FILLER_PATTERNS = [
    r"怎么办",
    r"怎么处理",
    r"怎么修",
    r"怎么查",
    r"怎么排查",
    r"如何处理",
    r"如何解决",
    r"什么原因",
    r"为什么",
    r"怎么回事",
    r"故障诊断",
    r"故障分析",
    r"维修思路",
    r"维修方案",
    r"维修指导",
    r"排查方法",
    r"解决方法",
]

NON_SEARCHABLE_TOPICS = {
    "根据数据流进行工况判断",
    "第三步输出维修指导",
}


@dataclass(frozen=True)
class RepairKnowledgeEntry:
    id: str
    title: str
    content: str
    topic: str
    normalized_topic: str
    searchable: bool
    title_category: str


class RepairKnowledgeService:
    """Load repair knowledge from Excel and expose title/context retrieval."""

    def __init__(self, excel_path: str | Path):
        path = Path(excel_path).expanduser()
        if not path.is_absolute():
            project_root = Path(__file__).resolve().parents[5]
            path = project_root / path
        self._excel_path = path.resolve()

    @classmethod
    def default_path(cls) -> Path:
        project_root = Path(__file__).resolve().parents[5]
        return project_root / "docs" / "fixdoc" / "维修知识库.xlsx"

    @classmethod
    def create_default(cls) -> "RepairKnowledgeService":
        return cls(cls.default_path())

    @cached_property
    def _entries(self) -> list[RepairKnowledgeEntry]:
        if not self._excel_path.exists():
            raise FileNotFoundError(f"Repair knowledge file not found: {self._excel_path}")

        workbook = load_workbook(self._excel_path, read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        entries: list[RepairKnowledgeEntry] = []
        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            title = str(row[0] or "").strip()
            content = str(row[1] or "").strip()
            if not title or not content:
                continue

            topic = self._extract_topic(title)
            normalized_topic = self._normalize_text(topic)
            title_category = self._classify_title(title, topic)
            searchable = bool(normalized_topic) and title_category in {"brand_specific", "generic_fault"}
            entries.append(
                RepairKnowledgeEntry(
                    id=f"repair_knowledge_{row_index}",
                    title=title,
                    content=content,
                    topic=topic,
                    normalized_topic=normalized_topic,
                    searchable=searchable,
                    title_category=title_category,
                )
            )
        return entries

    def lookup_titles(self, query: str) -> dict[str, Any]:
        normalized_query = self._normalize_query(query)
        scored_matches = self._match_entries(normalized_query)
        score_map = {entry.id: score for entry, score in scored_matches}

        searchable_entries = [entry for entry in self._entries if entry.searchable]
        ordered_entries = sorted(
            searchable_entries,
            key=lambda entry: (
                0 if entry.id in score_map else 1,
                -score_map.get(entry.id, 0.0),
                entry.topic,
                entry.title,
            ),
        )
        serialized_titles = [
            {
                "id": entry.id,
                "title": entry.title,
                "topic": entry.topic,
                "title_category": entry.title_category,
                "recall_score": round(score_map[entry.id], 2) if entry.id in score_map else None,
            }
            for entry in ordered_entries
        ]
        recommended_titles = [
            item
            for item in serialized_titles
            if item["recall_score"] is not None
        ][:8]

        return {
            "status": "ok",
            "data": {
                "query": query,
                "decision_mode": "llm_must_decide_match",
                "title_count": len(serialized_titles),
                "recommended_titles": recommended_titles,
                "titles": serialized_titles,
                "guidance": (
                    "These are title-level entries from the local repair knowledge Excel. "
                    "The local recall score only helps ordering and does not mean a confirmed match. "
                    "You must decide whether the user's issue really matches any title before loading full context."
                ),
            },
        }

    def load_context(self, entry_ids: list[str]) -> dict[str, Any]:
        entry_ids = [str(item).strip() for item in entry_ids if str(item).strip()]
        if not entry_ids:
            return {
                "status": "ok",
                "data": {
                    "loaded": False,
                    "entries": [],
                    "source_refs": [],
                    "primary_source": None,
                    "llm_context": "",
                },
            }

        by_id = {entry.id: entry for entry in self._entries if entry.searchable}
        selected_entries: list[RepairKnowledgeEntry] = []
        seen_ids: set[str] = set()
        for entry_id in entry_ids:
            entry = by_id.get(entry_id)
            if entry is None or entry.id in seen_ids:
                continue
            selected_entries.append(entry)
            seen_ids.add(entry.id)

        if not selected_entries:
            return {
                "status": "ok",
                "data": {
                    "loaded": False,
                    "entries": [],
                    "source_refs": [],
                    "primary_source": None,
                    "llm_context": "",
                },
            }

        selected_topics = {entry.normalized_topic for entry in selected_entries if entry.normalized_topic}
        related_entries: list[RepairKnowledgeEntry] = []
        for entry in self._entries:
            if not entry.searchable or entry.id in seen_ids:
                continue
            if entry.normalized_topic not in selected_topics:
                continue
            related_entries.append(entry)
            seen_ids.add(entry.id)

        bundled_entries = selected_entries + related_entries[:4]
        source_refs = []
        serialized_entries = []
        for index, entry in enumerate(bundled_entries):
            relation = "primary" if index == 0 else "related"
            source_refs.append(
                {
                    "id": entry.id,
                    "title": entry.title,
                    "relation": relation,
                    "match_score": 1.0 if entry in selected_entries else 0.8,
                }
            )
            serialized_entries.append(
                {
                    "id": entry.id,
                    "title": entry.title,
                    "topic": entry.topic,
                    "relation": relation,
                    "content": entry.content,
                }
            )

        return {
            "status": "ok",
            "data": {
                "loaded": True,
                "selected_entry_ids": [entry.id for entry in selected_entries],
                "source_refs": source_refs,
                "primary_source": source_refs[0],
                "entries": serialized_entries,
                "structured_evidence": self._extract_structured_evidence(serialized_entries),
                "llm_context": self._build_llm_context(serialized_entries),
            },
        }

    def lookup(self, query: str, supplemental_info: str | None = None) -> dict[str, Any]:
        """Backward-compatible alias for title lookup."""
        del supplemental_info
        return self.lookup_titles(query)

    def get_source_detail(self, entry_id: str) -> dict[str, Any] | None:
        entry = next((item for item in self._entries if item.id == entry_id), None)
        if entry is None:
            return None
        return {
            "id": entry.id,
            "title": entry.title,
            "topic": entry.topic,
            "content": entry.content,
        }

    @staticmethod
    def _extract_topic(title: str) -> str:
        topic = title.replace("*", "").strip()
        topic = re.sub(r"-标准版v?\d+$", "", topic, flags=re.IGNORECASE)
        topic = re.sub(r"标准版v?\d+$", "", topic, flags=re.IGNORECASE)
        for suffix in [
            "故障诊断分析提示词",
            "选择数据流的提示词",
            "能否定位故障点提示词",
            "定位故障点专用提示词",
            "定位故障点提示词",
            "通用提示词",
            "专用提示词",
            "提示词",
        ]:
            if topic.endswith(suffix):
                topic = topic[: -len(suffix)]
                break
        return topic.strip("：: -_[]（）()")

    @staticmethod
    def _classify_title(title: str, topic: str) -> str:
        if topic in NON_SEARCHABLE_TOPICS:
            return "workflow"
        if any(keyword in title for keyword in ["康明斯", "解放", "东风", "潍柴", "玉柴"]):
            return "brand_specific"
        if topic:
            return "generic_fault"
        return "workflow"

    @staticmethod
    def _normalize_text(value: str) -> str:
        if not value:
            return ""
        normalized = re.sub(r"[*#`_\s/（）()【】\[\]：:，,。.!！?？·\-]+", "", value)
        return normalized.lower()

    def _normalize_query(self, query: str) -> str:
        normalized = self._normalize_text(query)
        for pattern in QUESTION_FILLER_PATTERNS:
            normalized = re.sub(f"(?:{pattern})+$", "", normalized)
        normalized = re.sub(r"^(?:请问|帮我|麻烦|咨询一下|问一下|想问一下)", "", normalized)
        return normalized.strip()

    @staticmethod
    def _topic_bigrams(value: str) -> set[str]:
        if len(value) < 2:
            return {value} if value else set()
        return {value[index : index + 2] for index in range(len(value) - 1)}

    def _match_entries(self, normalized_query: str) -> list[tuple[RepairKnowledgeEntry, float]]:
        if not normalized_query:
            return []

        scored: list[tuple[RepairKnowledgeEntry, float]] = []
        for entry in self._entries:
            if not entry.searchable or not entry.normalized_topic:
                continue

            topic = entry.normalized_topic
            exact = topic == normalized_query
            contains = topic in normalized_query or normalized_query in topic
            bigram_hits = sum(1 for token in self._topic_bigrams(topic) if token and token in normalized_query)
            if not exact and not contains and bigram_hits == 0:
                continue

            score = 0.0
            if exact:
                score += 120.0
            if contains:
                score += 90.0 + min(len(topic), len(normalized_query))
            score += bigram_hits * 14.0
            score += len(set(topic) & set(normalized_query)) * 2.0
            if entry.title_category == "brand_specific":
                score += 10.0
            else:
                score -= 5.0

            scored.append((entry, score))

        scored.sort(key=lambda item: item[1], reverse=True)
        if not scored:
            return []

        primary_score = scored[0][1]
        return [
            item
            for item in scored
            if item[1] >= max(26.0, primary_score * 0.45)
        ]

    @staticmethod
    def _trim_llm_content(content: str, *, limit: int = 3200) -> str:
        cleaned = content.strip()
        if len(cleaned) <= limit:
            return cleaned
        return cleaned[:limit].rstrip() + "\n...[内容已截断]"

    def _build_llm_context(self, entries: list[dict[str, Any]]) -> str:
        lines = [
            "你已加载本地维修知识库正文，回答时优先参考这些资料。",
            "如果资料仍依赖缺失信息，请先调用 ask_user_question，问题由你结合资料内容自然组织，不要照搬固定模板。",
            "如果还缺少故障码、工况、数据流或 ECU 信息，不要直接写“还需补充”“请回复以下信息”“点击下方按钮”等文本，必须改为 ask_user_question。",
            "如果是启动/起动/打不着火/起动机无反应这类问题，生成 ask_user_question 时要优先给出可点选的预测候选项，尤其是现象、工况和报码方向，不能把这些字段留成空列表。",
            "不要写“由于缺乏针对性的维修案例”“当前证据不足”“资料不足”等会削弱用户信任的解释。",
            "如果缺信息但还不能明确列出要问什么，就直接给出当前最稳妥的排查建议，不要在正文末尾追加自由文本追问。",
            "面向用户回答时，不要暴露你的思考过程，不要写“根据维修经验，诊断的核心逻辑是……”这类元话术。",
            "不要把维修知识库内容直接翻译成一段标准答案，要把它转成现场排查逻辑。",
            "如果用户已经补充了现象、工况、报码或维修历史，你必须在答案里体现这些补充信息如何改变你的判断顺序。",
            "不要只给“检查A、检查B、检查C”这种并列列表，必须尽量写成“先查A，若正常再查B，若异常优先处理C”的顺序化路径。",
            "对于起动系统、供电系统、报码类问题，优先减少泛泛原理说明，增加结果导向的判断语句。",
            "如果你使用 markdown 标题作答，请直接从正文标题开始，不要先写“根据您的情况”“以下是建议”等开场过渡句。",
            "默认让“维修建议”成为篇幅最大的主体部分；“初步判断”和“优先检查”保持简洁，避免大段背景铺垫。",
        ]
        for index, entry in enumerate(entries, start=1):
            relation = "主参考" if entry.get("relation") == "primary" else "补充参考"
            lines.append(f"\n[维修经验{index}] {entry.get('title')}（{relation}）")
            if entry.get("topic"):
                lines.append(f"主题：{entry['topic']}")
            lines.append(self._trim_llm_content(str(entry.get("content") or "")))
        return "\n".join(lines)

    def _extract_structured_evidence(self, entries: list[dict[str, Any]]) -> dict[str, list[str]]:
        buckets: dict[str, list[str]] = {
            "checks": [],
            "thresholds": [],
            "actions": [],
            "recheck": [],
        }
        seen: dict[str, set[str]] = {key: set() for key in buckets}

        for entry in entries:
            content = str(entry.get("content") or "").strip()
            if not content:
                continue
            for raw_piece in re.split(r"[\n。；]", content):
                piece = re.sub(r"^\s*(?:#{1,6}\s*|[-*]|\d+[.、）)])\s*", "", raw_piece).strip()
                if not piece:
                    continue
                lowered = piece.lower()

                def _append(bucket: str, value: str) -> None:
                    normalized = value.lower()
                    if normalized in seen[bucket] or len(buckets[bucket]) >= 6:
                        return
                    seen[bucket].add(normalized)
                    buckets[bucket].append(value)

                if any(token in piece for token in ("检查", "测量", "量", "核对", "确认", "查看", "先看", "先查")):
                    _append("checks", piece)
                if (
                    re.search(r"\d", piece)
                    and any(unit in lowered for unit in ("v", "伏", "欧", "ohm", "%", "bar", "kpa", "rpm", "转"))
                ) or any(token in piece for token in ("正常", "接近", "偏高", "偏低", "在线", "离线", "压降")):
                    _append("thresholds", piece)
                if any(token in piece for token in ("修复", "更换", "处理", "清理", "隔离", "检修", "更正")):
                    _append("actions", piece)
                if any(token in piece for token in ("复验", "复测", "再次确认", "路试", "报码不再", "恢复正常")):
                    _append("recheck", piece)

        return {key: value for key, value in buckets.items() if value}
