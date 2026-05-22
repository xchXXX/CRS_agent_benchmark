#!/usr/bin/env python3
"""Randomly sample records from an Excel sheet into a new workbook."""

from __future__ import annotations

import argparse
import random
from pathlib import Path

from openpyxl import Workbook, load_workbook


DEFAULT_INPUT = Path("/Users/zhangjiexiang/Downloads/完整的电路图的目录.xlsx")
DEFAULT_OUTPUT = Path("/Users/zhangjiexiang/Downloads/完整的电路图的目录_随机150.xlsx")
DEFAULT_SAMPLE_SIZE = 150
DEFAULT_OUTPUT_SHEET = "随机抽样150"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="从 Excel 表格中随机抽取指定数量的数据记录。")
    parser.add_argument("--input", default=str(DEFAULT_INPUT), help="源 xlsx 文件路径")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT), help="输出 xlsx 文件路径")
    parser.add_argument("--sheet", default=None, help="源工作表名称；默认使用第一个工作表")
    parser.add_argument("--sample-size", type=int, default=DEFAULT_SAMPLE_SIZE, help="抽样记录数")
    parser.add_argument("--seed", type=int, default=None, help="随机种子；指定后可复现抽样结果")
    parser.add_argument("--output-sheet", default=DEFAULT_OUTPUT_SHEET, help="输出工作表名称")
    return parser.parse_args()


def copy_column_widths(source_ws, target_ws) -> None:
    for column_letter, dimension in source_ws.column_dimensions.items():
        if dimension.width:
            target_ws.column_dimensions[column_letter].width = dimension.width


def main() -> None:
    args = parse_args()
    input_path = Path(args.input).expanduser()
    output_path = Path(args.output).expanduser()

    if args.sample_size <= 0:
        raise ValueError("--sample-size 必须大于 0")
    if not input_path.exists():
        raise FileNotFoundError(f"源文件不存在: {input_path}")

    source_wb = load_workbook(input_path, read_only=False, data_only=False)
    source_ws = source_wb[args.sheet] if args.sheet else source_wb.worksheets[0]

    rows = list(source_ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("源工作表为空")

    header = rows[0]
    data_rows = rows[1:]
    if len(data_rows) < args.sample_size:
        raise ValueError(f"源工作表只有 {len(data_rows)} 条数据，不足以抽取 {args.sample_size} 条")

    rng = random.Random(args.seed)
    sampled_rows = rng.sample(data_rows, args.sample_size)

    output_wb = Workbook()
    output_ws = output_wb.active
    output_ws.title = args.output_sheet
    output_ws.append(header)
    for row in sampled_rows:
        output_ws.append(row)

    output_ws.freeze_panes = "A2"
    output_ws.auto_filter.ref = output_ws.dimensions
    copy_column_widths(source_ws, output_ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_wb.save(output_path)
    print(f"已从 {source_ws.title} 随机抽取 {args.sample_size} 条记录到: {output_path}")


if __name__ == "__main__":
    main()
