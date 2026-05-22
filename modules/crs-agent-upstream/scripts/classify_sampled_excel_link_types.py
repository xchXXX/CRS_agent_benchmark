#!/usr/bin/env python3
"""Classify sampled GGZJ document link types by exact first search result."""

from __future__ import annotations

import argparse
import os
import time
import unicodedata
from pathlib import Path
from typing import Any

import httpx
from openpyxl import load_workbook

from app.legacy.utils.token_utils import parse_jwt_source


DEFAULT_INPUT = Path("/Users/zhangjiexiang/Downloads/完整的电路图的目录_随机150.xlsx")
DEFAULT_OUTPUT = Path("/Users/zhangjiexiang/Downloads/完整的电路图的目录_随机150_链接类型.xlsx")
SEARCH_URL = "https://wx.51gonggui.com/commonrail/api/management/getMeansList.json"
VALIDATE_URL = "https://wx.51gonggui.com/commonrail/api/member-api/userLoginInfo.json"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="按 name 精确搜索并标注资料链接类型。")
    parser.add_argument("--input", default=str(DEFAULT_INPUT), help="源 xlsx 文件路径")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT), help="输出 xlsx 文件路径")
    parser.add_argument("--sheet", default=None, help="工作表名称；默认使用第一个工作表")
    parser.add_argument("--token-env", default="GGZJ_APP_TOKEN", help="读取 token 的环境变量名")
    parser.add_argument("--delay", type=float, default=0.8, help="每条搜索后的等待秒数")
    parser.add_argument("--page-size", type=int, default=5, help="每次搜索拉取的结果数")
    parser.add_argument("--limit", type=int, default=0, help="仅处理前 N 条数据；0 表示全部")
    return parser.parse_args()


def normalize_exact(value: Any) -> str:
    return unicodedata.normalize("NFKC", str(value or "")).strip()


def classify_data_type(item: dict[str, Any]) -> str:
    raw_value = item.get("dataType")
    try:
        data_type = int(raw_value)
    except (TypeError, ValueError):
        data_type = None

    if data_type == 2:
        return "wps"
    if data_type == 3:
        return "circuit"
    if data_type is not None:
        return f"legacy_{data_type}"
    return "unknown"


def validate_token(client: httpx.Client, token: str) -> None:
    response = client.post(
        VALIDATE_URL,
        json={},
        headers={
            "app-token": token,
            "source": parse_jwt_source(token),
            "Content-Type": "application/json",
        },
    )
    response.raise_for_status()
    data = response.json()
    if not (
        isinstance(data, dict)
        and data.get("status") == 200
        and isinstance(data.get("data"), dict)
        and data["data"].get("userId")
    ):
        message = data.get("msg", "token 无效") if isinstance(data, dict) else "token 无效"
        raise RuntimeError(message)


def search_first(client: httpx.Client, *, token: str, query: str, page_size: int) -> dict[str, Any] | None:
    response = client.post(
        SEARCH_URL,
        json={
            "parentClassId": "0",
            "dataName": query,
            "pageNum": 1,
            "pageSize": page_size,
        },
        headers={
            "Content-Type": "application/json;charset=UTF-8",
            "app-token": token,
            "source": parse_jwt_source(token),
        },
    )
    response.raise_for_status()
    data = response.json()
    if data.get("status") != 200:
        raise RuntimeError(data.get("msg", "搜索接口返回异常"))
    items = data.get("data", {}).get("dataList", [])
    return items[0] if items else None


def main() -> None:
    args = parse_args()
    token = os.getenv(args.token_env, "").strip()
    if not token:
        raise RuntimeError(f"请先设置环境变量 {args.token_env}")

    input_path = Path(args.input).expanduser()
    output_path = Path(args.output).expanduser()
    workbook = load_workbook(input_path)
    worksheet = workbook[args.sheet] if args.sheet else workbook.worksheets[0]

    headers = [cell.value for cell in worksheet[1]]
    if "name" not in headers:
        raise RuntimeError("源表缺少 name 列")
    name_col = headers.index("name") + 1

    link_type_col = len(headers) + 1
    worksheet.cell(row=1, column=link_type_col, value="link_type")

    rows_to_process = worksheet.max_row - 1
    if args.limit > 0:
        rows_to_process = min(rows_to_process, args.limit)

    timeout = httpx.Timeout(20.0, connect=10.0)
    with httpx.Client(timeout=timeout) as client:
        validate_token(client, token)
        for offset, row_idx in enumerate(range(2, 2 + rows_to_process), start=1):
            name = normalize_exact(worksheet.cell(row=row_idx, column=name_col).value)
            if not name:
                link_type = "empty_name"
            else:
                try:
                    first = search_first(client, token=token, query=name, page_size=args.page_size)
                    if first is None:
                        link_type = "no_results"
                    else:
                        first_name = normalize_exact(first.get("dataNameWs") or first.get("dataName"))
                        if first_name == name:
                            link_type = classify_data_type(first)
                        else:
                            link_type = "not_exact_match"
                except Exception as exc:
                    link_type = f"error: {exc}"

            worksheet.cell(row=row_idx, column=link_type_col, value=link_type)
            print(f"[{offset}/{rows_to_process}] row={row_idx} link_type={link_type} name={name[:60]}")
            if offset < rows_to_process and args.delay > 0:
                time.sleep(args.delay)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    print(f"已输出: {output_path}")


if __name__ == "__main__":
    main()
